from __future__ import annotations

import json
import sqlite3
from datetime import UTC, datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer

import threading
import time

from services.terminal_dashboard import build_dashboard_snapshot, render_terminal_dashboard
from execution.execution_health import read_execution_health
from storage.execution_queue_store import ExecutionQueueStore
from storage.account_store import AccountStore
from storage.fund_store import FundStore
from services.line_bot import handle_webhook, notify_new_account
from services.web_spa import build_spa_html
import os

# Rate limiting for admin login
_login_attempts: dict = {}
_login_lock = threading.Lock()


def _get_client_ip(handler) -> str:
    forwarded = handler.headers.get("X-Forwarded-For", "")
    return forwarded.split(",")[0].strip() if forwarded else handler.client_address[0]


def _is_rate_limited(ip: str) -> tuple[bool, int]:
    with _login_lock:
        rec = _login_attempts.get(ip, {})
        locked_until = rec.get("locked_until", 0)
        if locked_until > time.time():
            return True, int(locked_until - time.time())
        return False, 0


def _record_fail(ip: str) -> None:
    with _login_lock:
        rec = _login_attempts.setdefault(ip, {"count": 0, "locked_until": 0})
        rec["count"] += 1
        if rec["count"] >= 5:
            rec["locked_until"] = time.time() + 900  # 15 min
            rec["count"] = 0


def _record_success(ip: str) -> None:
    with _login_lock:
        _login_attempts.pop(ip, None)


_COUNTABLE_REASONS = {"TAKE_PROFIT_1", "TAKE_PROFIT_2", "TAKE_PROFIT_3", "STOP_LOSS"}
_TP_W1, _TP_W2, _TP_W3 = 0.40, 0.30, 0.30  # must match execution/models.py defaults


def _realized_rr(row) -> float:
    """Replicate google_sheets_sync._weighted_realized_rr for closed STOPPED trades."""
    reason = (row["close_reason"] or "").upper()
    rr1 = float(row["rr_tp1"]) if row["rr_tp1"] else None
    rr2 = float(row["rr_tp2"]) if row["rr_tp2"] else None
    rr3 = float(row["rr_tp3"]) if row["rr_tp3"] else None
    tp1_hit = bool(row["tp1_hit_at"])
    tp2_hit = bool(row["tp2_hit_at"])
    tp3_hit = bool(row["tp3_hit_at"])

    rr = 0.0
    if tp1_hit and rr1 is not None:
        rr += _TP_W1 * rr1
    if tp2_hit and rr2 is not None:
        rr += _TP_W2 * rr2
    if tp3_hit and rr3 is not None:
        rr += _TP_W3 * rr3

    # residual position that wasn't closed at a TP
    remaining = 1.0 - ((_TP_W1 if tp1_hit else 0) + (_TP_W2 if tp2_hit else 0) + (_TP_W3 if tp3_hit else 0))
    remaining = max(remaining, 0.0)
    if remaining > 0 and reason == "STOP_LOSS":
        stop = float(row["managed_stop_loss"]) if row["managed_stop_loss"] else float(row["stop_loss"])
        entry = float(row["entry_triggered_price"] or row["entry_price"])
        sl_original = float(row["stop_loss"])
        risk = abs(entry - sl_original)
        if risk > 0:
            side = (row["side"] or "").upper()
            if side == "LONG":
                residual = (stop - entry) / risk
            else:
                residual = (entry - stop) / risk
            rr += remaining * residual
        else:
            rr += remaining * -1.0
    elif remaining > 0:
        rr += remaining * -1.0

    return round(rr, 4)


def _trade_result(row) -> tuple[str, float]:
    """Returns (result_label, realized_rr). Only call for countable trades."""
    reason = (row["close_reason"] or "").upper()
    tp1_hit = bool(row["tp1_hit_at"])
    tp2_hit = bool(row["tp2_hit_at"])
    tp3_hit = bool(row["tp3_hit_at"])
    rr = _realized_rr(row)

    if reason == "TAKE_PROFIT_3" or tp3_hit:
        return "TP3_HIT", rr
    if reason == "STOP_LOSS":
        if tp2_hit:
            return "TP2_THEN_SL", rr
        if tp1_hit:
            return "TP1_THEN_SL", rr
        return "SL_HIT", rr
    return reason, rr


def _build_trade_stats(db_path: str) -> dict:
    try:
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("""
            SELECT close_reason, side, entry_price, entry_triggered_price, stop_loss,
                   managed_stop_loss, rr_tp1, rr_tp2, rr_tp3,
                   tp1_hit_at, tp2_hit_at, tp3_hit_at
            FROM signals
            WHERE close_reason IN ('TAKE_PROFIT_1','TAKE_PROFIT_2','TAKE_PROFIT_3','STOP_LOSS')
              AND entry_triggered_at IS NOT NULL
            ORDER BY closed_at ASC
        """)
        rows = cur.fetchall()
        conn.close()
    except Exception:
        return {"win_rate": 0, "avg_rr": 0, "total": 0, "wins": 0, "losses": 0, "profit_factor": 0, "max_dd": 0}

    wins = losses = 0
    rr_list: list[float] = []
    equity = 0.0
    peak = 0.0
    max_dd = 0.0

    for row in rows:
        _label, rr = _trade_result(row)
        if rr > 0:
            wins += 1
        else:
            losses += 1
        rr_list.append(rr)
        equity += rr
        peak = max(peak, equity)
        max_dd = max(max_dd, peak - equity)

    total = wins + losses
    win_rate = round(wins / total * 100, 1) if total > 0 else 0
    avg_rr = round(sum(rr_list) / len(rr_list), 2) if rr_list else 0
    gross_win = sum(r for r in rr_list if r > 0)
    gross_loss = abs(sum(r for r in rr_list if r < 0))
    profit_factor = round(gross_win / gross_loss, 2) if gross_loss > 0 else 0

    return {
        "win_rate": win_rate,
        "avg_rr": avg_rr,
        "total": total,
        "wins": wins,
        "losses": losses,
        "profit_factor": profit_factor,
        "max_dd": round(max_dd, 2),
    }


def _build_trade_history(db_path: str, limit: int = 200) -> list[dict]:
    try:
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("""
            SELECT symbol, timeframe, side, entry_price, stop_loss, managed_stop_loss,
                   close_reason, entry_triggered_price,
                   tp1_hit_at, tp2_hit_at, tp3_hit_at,
                   tp1_hit_price, tp2_hit_price, tp3_hit_price,
                   rr_tp1, rr_tp2, rr_tp3,
                   closed_at, created_at
            FROM signals
            WHERE close_reason IN ('TAKE_PROFIT_1','TAKE_PROFIT_2','TAKE_PROFIT_3','STOP_LOSS')
              AND entry_triggered_at IS NOT NULL
            ORDER BY closed_at DESC
            LIMIT ?
        """, (limit,))
        rows = cur.fetchall()
        conn.close()
    except Exception:
        return []

    result = []
    for row in rows:
        label, rr = _trade_result(row)
        is_win = rr > 0
        result.append({
            "closed_at": (row["closed_at"] or row["created_at"] or "")[:16].replace("T", " "),
            "symbol": row["symbol"],
            "timeframe": row["timeframe"],
            "side": (row["side"] or "").upper(),
            "entry": row["entry_price"],
            "exit": row["tp3_hit_price"] or row["tp2_hit_price"] or row["tp1_hit_price"] or row["entry_triggered_price"],
            "close_reason": label,
            "rr": round(rr, 2),
            "result": "WIN" if is_win else "LOSS",
        })
    return result


def _build_active_trades(db_path: str) -> list[dict]:
    try:
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("""
            SELECT symbol, timeframe, side, status, entry_price, entry_triggered_price,
                   stop_loss, tp1, tp2, tp3, tp1_hit_at, tp2_hit_at, created_at
            FROM signals
            WHERE status IN ('ACTIVE','PARTIAL_TP1','PARTIAL_TP2')
              AND entry_triggered_at IS NOT NULL
            ORDER BY created_at DESC
        """)
        rows = cur.fetchall()
        conn.close()
    except Exception:
        return []

    result = []
    for row in rows:
        result.append({
            "symbol": row["symbol"],
            "timeframe": row["timeframe"],
            "side": (row["side"] or "").upper(),
            "status": row["status"],
            "entry": row["entry_triggered_price"] or row["entry_price"],
            "sl": row["stop_loss"],
            "tp1": row["tp1"],
            "tp2": row["tp2"],
            "tp3": row["tp3"],
            "tp1_hit": bool(row["tp1_hit_at"]),
            "tp2_hit": bool(row["tp2_hit_at"]),
        })
    return result


_SHARED_CSS = """
* { box-sizing: border-box; margin: 0; padding: 0; }
:root {
  --bg: #0a0e1a; --surface: #111827; --surface-2: #1f2937;
  --border: #1f2937; --text: #e5e7eb; --muted: #9ca3af;
  --primary: #3b82f6; --success: #10b981; --danger: #ef4444; --warning: #f59e0b;
  --radius: 12px;
}
body.light {
  --bg: #f1f5f9; --surface: #ffffff; --surface-2: #e2e8f0;
  --border: #e2e8f0; --text: #0f172a; --muted: #64748b;
}
body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; background: var(--bg); color: var(--text); min-height: 100vh; padding: 16px; transition: background .2s, color .2s; }
.container { max-width: 1200px; margin: 0 auto; }
.header { display: flex; justify-content: space-between; align-items: center; padding: 16px 20px; background: var(--surface); border-radius: var(--radius); margin-bottom: 16px; border: 1px solid var(--border); }
.header h1 { font-size: 20px; font-weight: 700; }
.nav { display: flex; gap: 8px; }
.nav a { padding: 8px 16px; border-radius: 8px; color: var(--muted); text-decoration: none; font-size: 14px; font-weight: 500; transition: .15s; }
.nav a:hover { background: var(--surface-2); color: var(--text); }
.nav a.active { background: var(--primary); color: #fff; }
.stats-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(160px, 1fr)); gap: 12px; margin-bottom: 16px; }
.stat { background: var(--surface); border: 1px solid var(--border); border-radius: var(--radius); padding: 16px; }
.stat-label { font-size: 11px; color: var(--muted); text-transform: uppercase; letter-spacing: .5px; margin-bottom: 6px; }
.stat-value { font-size: 24px; font-weight: 700; }
.stat-sub { font-size: 12px; color: var(--muted); margin-top: 4px; }
.text-success { color: var(--success); }
.text-danger { color: var(--danger); }
.text-warning { color: var(--warning); }
.section { background: var(--surface); border: 1px solid var(--border); border-radius: var(--radius); padding: 20px; margin-bottom: 16px; }
.section-head { display: flex; justify-content: space-between; align-items: center; margin-bottom: 16px; flex-wrap: wrap; gap: 12px; }
.section h2 { font-size: 16px; font-weight: 600; }
.filter-chips { display: flex; gap: 6px; }
.chip { padding: 6px 12px; border-radius: 6px; background: var(--surface-2); color: var(--muted); font-size: 13px; cursor: pointer; border: 1px solid transparent; }
.chip.active { background: var(--primary); color: #fff; }
.table { width: 100%; border-collapse: collapse; font-size: 13px; }
.table th, .table td { padding: 10px 12px; text-align: left; border-bottom: 1px solid var(--border); }
.table th { color: var(--muted); font-weight: 500; font-size: 11px; text-transform: uppercase; letter-spacing: .5px; }
.table tr:last-child td { border-bottom: none; }
.badge { padding: 3px 8px; border-radius: 4px; font-size: 11px; font-weight: 600; display: inline-block; }
.badge-long { background: rgba(16,185,129,.15); color: var(--success); }
.badge-short { background: rgba(239,68,68,.15); color: var(--danger); }
.badge-win { background: rgba(16,185,129,.15); color: var(--success); }
.badge-loss { background: rgba(239,68,68,.15); color: var(--danger); }
.badge-bull { background: rgba(16,185,129,.15); color: var(--success); }
.badge-bear { background: rgba(239,68,68,.15); color: var(--danger); }
.empty { text-align: center; padding: 40px; color: var(--muted); font-size: 14px; }
.table-wrap { overflow-x: auto; }
.err-banner { background: rgba(239,68,68,.1); border: 1px solid rgba(239,68,68,.3); border-radius: 8px; padding: 10px 14px; color: var(--danger); font-size: 13px; margin-bottom: 12px; }
.pos { color: var(--success); font-weight: 600; }
.neg { color: var(--danger); font-weight: 600; }
.muted { color: var(--muted); }
.api-card { background: var(--surface); border: 1px solid var(--border); border-radius: var(--radius); padding: 20px; margin-bottom: 16px; }
.api-card h2 { font-size: 14px; color: var(--muted); margin-bottom: 12px; text-transform: uppercase; letter-spacing: .5px; }
.api-row { display: grid; grid-template-columns: 1fr 1fr auto auto; gap: 8px; align-items: end; }
.api-row input { background: var(--bg); border: 1px solid var(--border); border-radius: 8px; padding: 10px 12px; color: var(--text); font-size: 14px; font-family: monospace; width: 100%; }
.api-row input:focus { outline: none; border-color: var(--primary); }
.api-status { font-size: 12px; color: var(--muted); margin-top: 8px; }
.api-status .ok { color: var(--success); font-weight: 600; }
.api-status .err { color: var(--danger); font-weight: 600; }
.btn { padding: 10px 16px; border-radius: 8px; border: none; cursor: pointer; font-size: 14px; font-weight: 600; transition: .15s; white-space: nowrap; }
.btn-primary { background: var(--primary); color: #fff; }
.btn-primary:hover { background: #2563eb; }
.btn-ghost { background: var(--surface-2); color: var(--muted); }
.btn-ghost:hover { color: var(--text); }
.nav-guide { padding: 8px 14px; border-radius: 8px; background: rgba(245,158,11,.15); color: var(--warning); font-size: 13px; font-weight: 600; text-decoration: none; border: 1px solid rgba(245,158,11,.3); white-space: nowrap; }
.nav-theme { padding: 8px 14px; border-radius: 8px; border: 1px solid var(--primary); background: transparent; color: var(--primary); font-size: 13px; font-weight: 600; cursor: pointer; white-space: nowrap; }
@media (max-width: 640px) {
  body { padding: 8px; }
  .header { flex-direction: column; gap: 8px; padding: 12px 14px; overflow: hidden; }
  .header h1 { font-size: 16px; }
  .header > div:first-child > div { font-size: 10px; }
  .nav { overflow-x: auto; -webkit-overflow-scrolling: touch; flex-wrap: nowrap; gap: 5px; padding-bottom: 4px; width: 100%; }
  .nav a, .nav-guide, .nav-theme { font-size: 11px !important; padding: 6px 10px !important; white-space: nowrap; }
  .stats-grid { grid-template-columns: 1fr 1fr; gap: 8px; }
  .stat { padding: 12px; }
  .stat-value { font-size: 20px; }
  .stat-label { font-size: 10px; }
  .section { padding: 14px 12px; }
  .section h2 { font-size: 14px; }
  .table th, .table td { padding: 8px 6px; font-size: 12px; }
  .api-card { padding: 14px; }
  .api-row { grid-template-columns: 1fr 1fr; }
  .api-row input { grid-column: 1 / -1; }
  .section-head { flex-direction: column; align-items: flex-start; gap: 8px; }
}
"""


def build_web_dashboard_html(symbol: str, refresh_seconds: float) -> str:
    refresh_ms = max(int(refresh_seconds * 1000), 1000)
    escaped_symbol = json.dumps(symbol)
    return f"""<!DOCTYPE html>
<html lang="en"><head>
<meta charset="UTF-8" /><meta name="viewport" content="width=device-width, initial-scale=1" />
<title>Trading Dashboard</title>
<style>{_SHARED_CSS}</style>
</head><body>
<div class="container">
  <div class="header">
    <div>
      <h1>&#x26A1; Trading Dashboard</h1>
      <div style="font-size:11px;color:var(--muted);margin-top:4px;letter-spacing:.3px;">Since <span style="color:var(--primary);font-weight:600;">18 Mar 2026</span> &nbsp;&bull;&nbsp; Elliott Wave Engine &nbsp;&bull;&nbsp; &#x1F441; <span id="visit-count" style="color:var(--text);font-weight:600;">...</span> views</div>
    </div>
    <nav class="nav">
      <a href="/" class="active">Dashboard</a>
      <a href="/history">History</a>
      <a href="/board">&#x1F4CB; Board</a>
      <a href="/register" style="padding:8px 14px;border-radius:8px;background:rgba(16,185,129,.15);color:var(--success);font-size:13px;font-weight:600;text-decoration:none;border:1px solid rgba(16,185,129,.3);white-space:nowrap;">&#x270F; Join</a>
      <a href="/login" style="padding:8px 14px;border-radius:8px;background:rgba(59,130,246,.15);color:var(--primary);font-size:13px;font-weight:600;text-decoration:none;border:1px solid rgba(59,130,246,.3);white-space:nowrap;">&#x1F511; Login</a>
      <a href="/guide" class="nav-guide">&#x1F4D6; API Guide</a>
      <button onclick="toggleTheme()" id="theme-btn" class="nav-theme">&#x263D; Dark</button>
    </nav>
  </div>
  <div id="error-slot"></div>

  <div class="stats-grid">
    <div class="stat"><div class="stat-label">Win Rate</div><div class="stat-value" id="s-wr">&ndash;</div><div class="stat-sub" id="s-wl">&ndash;</div></div>
    <div class="stat"><div class="stat-label">Avg RR</div><div class="stat-value" id="s-rr">&ndash;</div><div class="stat-sub">risk:reward</div></div>
    <div class="stat"><div class="stat-label">Max Drawdown</div><div class="stat-value text-danger" id="s-dd">&ndash;</div><div class="stat-sub">in R units</div></div>
    <div class="stat"><div class="stat-label">Total Trades</div><div class="stat-value" id="s-total">&ndash;</div><div class="stat-sub">closed</div></div>
    <div class="stat"><div class="stat-label">Profit Factor</div><div class="stat-value" id="s-pf">&ndash;</div><div class="stat-sub">gross W/L</div></div>
  </div>
  <div class="section">
    <div class="section-head">
      <h2>&#x1F4CC; Open Trades <span id="active-count" style="color:var(--muted);font-weight:400;">(0)</span></h2>
    </div>
    <div class="table-wrap">
      <table class="table">
        <thead><tr><th>Symbol</th><th>TF</th><th>Side</th><th>Entry</th><th>SL</th><th>TP1</th><th>TP2</th><th>TP3</th><th>Status</th></tr></thead>
        <tbody id="active-body"><tr><td colspan="9" class="empty">Loading&hellip;</td></tr></tbody>
      </table>
    </div>
  </div>
</div>
<script>
(function(){{
  const SYMBOL = {escaped_symbol};
  const REFRESH_MS = {refresh_ms};

  // Dark/Light theme toggle
  function applyTheme(mode) {{
    document.body.classList.toggle("light", mode === "light");
    const btn = document.getElementById("theme-btn");
    if (btn) btn.textContent = mode === "light" ? "\U0001F319 Dark Mode" : "\u2600 Light Mode";
  }}
  function toggleTheme() {{
    const next = document.body.classList.contains("light") ? "dark" : "light";
    localStorage.setItem("ew_theme", next);
    applyTheme(next);
  }}
  window.toggleTheme = toggleTheme;
  applyTheme(localStorage.getItem("ew_theme") || "dark");

  function fmt(n, d) {{
    if (n == null || n === "" || isNaN(Number(n))) return "\u2013";
    return Number(n).toLocaleString("en-US", {{minimumFractionDigits: d, maximumFractionDigits: d}});
  }}

  async function tick() {{
    try {{
      const r = await fetch("/api/snapshot", {{cache: "no-store"}});
      if (!r.ok) throw new Error("HTTP " + r.status);
      const d = await r.json();
      if (!d.ok) throw new Error(d.error || "snapshot failed");
      render(d.snapshot);
      document.getElementById("error-slot").innerHTML = "";
    }} catch(e) {{
      document.getElementById("error-slot").innerHTML =
        '<div class="err-banner">Error: ' + (e.message||e) + '</div>';
    }}
  }}

  function render(d) {{
    const st = d.stats || {{}};
    const wr = st.win_rate ?? null;
    const wrEl = document.getElementById("s-wr");
    wrEl.textContent = wr != null ? wr + "%" : "\u2013";
    wrEl.className = "stat-value" + (wr >= 50 ? " text-success" : wr != null ? " text-danger" : "");
    document.getElementById("s-wl").textContent = st.total ? st.wins + "W / " + st.losses + "L of " + st.total : "\u2013";
    document.getElementById("s-rr").textContent = st.avg_rr ?? "\u2013";
    document.getElementById("s-dd").textContent = st.max_dd != null ? "-" + st.max_dd + "R" : "\u2013";
    document.getElementById("s-total").textContent = st.total ?? "\u2013";
    document.getElementById("s-pf").textContent = st.profit_factor ?? "\u2013";

    const acts = Array.isArray(d.active_trades) ? d.active_trades : [];
    document.getElementById("active-count").textContent = "(" + acts.length + ")";
    document.getElementById("active-body").innerHTML = acts.length
      ? acts.map(t => {{
          const side = (t.side||"").toUpperCase();
          const sideBadge = side === "LONG"
            ? "<span class='badge badge-long'>LONG</span>"
            : "<span class='badge badge-short'>SHORT</span>";
          const status = (t.status||"").replace("_"," ");
          return "<tr>"
            + "<td><b>" + (t.symbol||"\u2013") + "</b></td>"
            + "<td>" + (t.timeframe||"\u2013") + "</td>"
            + "<td>" + sideBadge + "</td>"
            + "<td>" + fmt(t.entry,4) + "</td>"
            + "<td class='neg'>" + fmt(t.sl,4) + "</td>"
            + "<td class='pos'>" + fmt(t.tp1,4) + "</td>"
            + "<td class='pos'>" + fmt(t.tp2,4) + "</td>"
            + "<td class='pos'>" + fmt(t.tp3,4) + "</td>"
            + "<td><span class='badge'>" + status + "</span></td>"
            + "</tr>";
        }}).join("")
      : "<tr><td colspan='9' class='empty'>No open trades</td></tr>";
  }}

  tick();
  setInterval(tick, REFRESH_MS);

  fetch("/api/visits").then(r=>r.json()).then(d=>{{
    const el = document.getElementById("visit-count");
    if (el && d.ok) el.textContent = Number(d.visits).toLocaleString();
  }}).catch(()=>{{}});
}})();
</script>
</body></html>
"""


def build_history_html(refresh_seconds: float) -> str:
    refresh_ms = max(int(refresh_seconds * 1000), 1000)
    return f"""<!DOCTYPE html>
<html lang="en"><head>
<meta charset="UTF-8" /><meta name="viewport" content="width=device-width, initial-scale=1" />
<title>Trade History</title>
<style>{_SHARED_CSS}</style>
</head><body>
<div class="container">
  <div class="header">
    <h1>&#x26A1; Trading Dashboard</h1>
    <nav class="nav">
      <a href="/">Dashboard</a>
      <a href="/history" class="active">History</a>
      <button onclick="toggleTheme()" id="theme-btn" class="nav-theme">&#x263D; Dark</button>
    </nav>
  </div>
  <div id="error-slot"></div>
  <div class="section">
    <div class="section-head">
      <h2>&#x1F4C5; Trade History</h2>
      <div class="filter-chips">
        <span class="chip" onclick="setFilter('7')">7 Days</span>
        <span class="chip active" id="chip-30" onclick="setFilter('30')">30 Days</span>
        <span class="chip" onclick="setFilter('all')">All</span>
      </div>
    </div>
    <div class="table-wrap">
      <table class="table">
        <thead><tr><th>Closed</th><th>Symbol</th><th>TF</th><th>Side</th><th>Entry</th><th>Exit</th><th>RR</th><th>Result</th></tr></thead>
        <tbody id="history-body"><tr><td colspan="8" class="empty">Loading&hellip;</td></tr></tbody>
      </table>
    </div>
  </div>
</div>
<script>
(function(){{
  const REFRESH_MS = {refresh_ms};
  let currentFilter = "30";
  let allTrades = [];

  function applyTheme(mode) {{
    document.body.classList.toggle("light", mode === "light");
    const btn = document.getElementById("theme-btn");
    if (btn) btn.textContent = mode === "light" ? "\U0001F319 Dark Mode" : "\u2600 Light Mode";
  }}
  function toggleTheme() {{
    const next = document.body.classList.contains("light") ? "dark" : "light";
    localStorage.setItem("ew_theme", next);
    applyTheme(next);
  }}
  window.toggleTheme = toggleTheme;
  applyTheme(localStorage.getItem("ew_theme") || "dark");

  function setFilter(f) {{
    currentFilter = f;
    document.querySelectorAll(".chip").forEach(c => c.classList.remove("active"));
    event.target.classList.add("active");
    renderTable();
  }}
  window.setFilter = setFilter;

  function filterTrades(trades, f) {{
    if (f === "all") return trades;
    const days = parseInt(f);
    const cutoff = new Date();
    cutoff.setDate(cutoff.getDate() - days);
    return trades.filter(t => t.closed_at && new Date(t.closed_at) >= cutoff);
  }}

  function fmt(n, d) {{
    if (n == null || n === "" || isNaN(Number(n))) return "\u2013";
    return Number(n).toLocaleString("en-US", {{minimumFractionDigits: d, maximumFractionDigits: d}});
  }}

  function renderTable() {{
    const trades = filterTrades(allTrades, currentFilter);
    const tbody = document.getElementById("history-body");
    if (!trades.length) {{
      tbody.innerHTML = "<tr><td colspan='8' class='empty'>No trades found</td></tr>";
      return;
    }}
    tbody.innerHTML = trades.map(t => {{
      const side = (t.side||"").toUpperCase();
      const isWin = t.result === "WIN";
      return "<tr>"
        + "<td class='muted'>" + (t.closed_at||"\u2013") + "</td>"
        + "<td><b>" + (t.symbol||"\u2013") + "</b></td>"
        + "<td class='muted'>" + (t.timeframe||"\u2013") + "</td>"
        + "<td><span class='badge " + (side==="LONG"?"badge-long":"badge-short") + "'>" + side + "</span></td>"
        + "<td>" + fmt(t.entry,4) + "</td>"
        + "<td>" + fmt(t.exit,4) + "</td>"
        + "<td class='" + (t.rr>0?"pos":"neg") + "'>" + (t.rr>0?"+":"") + t.rr + "R</td>"
        + "<td><span class='badge " + (isWin?"badge-win":"badge-loss") + "'>" + (t.result||"\u2013") + "</span></td>"
        + "</tr>";
    }}).join("");
  }}

  async function tick() {{
    try {{
      const r = await fetch("/api/history", {{cache: "no-store"}});
      if (!r.ok) throw new Error("HTTP " + r.status);
      const d = await r.json();
      if (!d.ok) throw new Error(d.error || "history failed");
      allTrades = d.trades || [];
      renderTable();
      document.getElementById("error-slot").innerHTML = "";
    }} catch(e) {{
      document.getElementById("error-slot").innerHTML =
        '<div class="err-banner">Error: ' + (e.message||e) + '</div>';
    }}
  }}

  tick();
  setInterval(tick, REFRESH_MS);
}})();
</script>
</body></html>
"""


_ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "admin1234")


def _ensure_posts_table(db_path: str) -> None:
    conn = sqlite3.connect(db_path, timeout=30)
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS posts (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            name       TEXT NOT NULL,
            message    TEXT NOT NULL,
            created_at TEXT NOT NULL
        )
    """)
    conn.commit()
    conn.close()


def _get_posts(db_path: str, limit: int = 100) -> list[dict]:
    try:
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT id, name, message, created_at FROM posts ORDER BY id DESC LIMIT ?", (limit,)
        ).fetchall()
        conn.close()
        return [{"id": r["id"], "name": r["name"], "message": r["message"],
                 "created_at": (r["created_at"] or "")[:16].replace("T", " ")} for r in rows]
    except Exception:
        return []


def _create_post(db_path: str, name: str, message: str) -> int:
    from datetime import UTC, datetime
    now = datetime.now(UTC).replace(microsecond=0).isoformat()
    conn = sqlite3.connect(db_path, timeout=30)
    conn.execute("PRAGMA journal_mode=WAL;")
    cur = conn.execute(
        "INSERT INTO posts (name, message, created_at) VALUES (?, ?, ?)", (name, message, now)
    )
    post_id = cur.lastrowid
    conn.commit()
    conn.close()
    return post_id


def build_board_html() -> str:
    return f"""<!DOCTYPE html>
<html lang="en"><head>
<meta charset="UTF-8" /><meta name="viewport" content="width=device-width, initial-scale=1" />
<title>Message Board — AlphaFutures</title>
<style>{_SHARED_CSS}
.board-wrap {{ max-width: 720px; margin: 0 auto; }}
.post-form {{ background: var(--surface); border: 1px solid var(--border); border-radius: var(--radius); padding: 24px; margin-bottom: 20px; }}
.post-form h2 {{ font-size: 16px; font-weight: 700; margin-bottom: 16px; }}
.field {{ margin-bottom: 14px; }}
.field label {{ display: block; font-size: 11px; color: var(--muted); text-transform: uppercase; letter-spacing: .5px; margin-bottom: 6px; }}
.field input, .field textarea {{ width: 100%; background: var(--bg); border: 1px solid var(--border); border-radius: 8px; padding: 10px 12px; color: var(--text); font-size: 14px; font-family: inherit; outline: none; }}
.field textarea {{ min-height: 100px; resize: vertical; line-height: 1.6; }}
.field input:focus, .field textarea:focus {{ border-color: var(--primary); }}
.post-card {{ background: var(--surface); border: 1px solid var(--border); border-radius: var(--radius); padding: 18px 20px; margin-bottom: 10px; }}
.post-header {{ display: flex; justify-content: space-between; align-items: center; margin-bottom: 10px; gap: 12px; flex-wrap: wrap; }}
.post-name {{ font-weight: 700; font-size: 14px; }}
.post-time {{ font-size: 11px; color: var(--muted); }}
.post-msg {{ font-size: 14px; color: var(--text); line-height: 1.7; white-space: pre-wrap; word-break: break-word; }}
.submit-btn {{ padding: 10px 24px; background: var(--primary); color: #fff; border: none; border-radius: 8px; font-size: 14px; font-weight: 700; cursor: pointer; }}
.submit-btn:disabled {{ opacity: .5; cursor: not-allowed; }}
.empty-board {{ text-align: center; padding: 48px; color: var(--muted); font-size: 14px; }}
.msg-ok {{ background: rgba(16,185,129,.12); color: var(--success); padding: 10px 14px; border-radius: 8px; font-size: 13px; margin-top: 10px; display: none; }}
.msg-err {{ background: rgba(239,68,68,.12); color: var(--danger); padding: 10px 14px; border-radius: 8px; font-size: 13px; margin-top: 10px; display: none; }}
</style></head><body>
<div class="container">
  <div class="header">
    <div>
      <h1>&#x26A1; AlphaFutures</h1>
      <div style="font-size:11px;color:var(--muted);margin-top:4px;">Message Board — Issues &amp; Feedback</div>
    </div>
    <nav class="nav">
      <a href="/">Dashboard</a>
      <a href="/history">History</a>
      <a href="/board" class="active">&#x1F4CB; Board</a>
    </nav>
  </div>

  <div class="board-wrap">
    <div class="post-form">
      <h2>&#x270F; Submit Issue / Feedback</h2>
      <div class="field">
        <label>Your Name</label>
        <input type="text" id="post-name" placeholder="Name or nickname" maxlength="60" />
      </div>
      <div class="field">
        <label>Message</label>
        <textarea id="post-msg" placeholder="Describe your issue or feedback here..." maxlength="2000"></textarea>
      </div>
      <button class="submit-btn" id="submit-btn" onclick="submitPost()">&#x1F4E8; Submit</button>
      <div class="msg-ok" id="msg-ok">&#x2713; Message sent. Thank you!</div>
      <div class="msg-err" id="msg-err"></div>
    </div>

    <div id="posts-list"><p class="empty-board">&#x23F3; Loading...</p></div>
  </div>
</div>
<script>
async function submitPost() {{
  const name = document.getElementById("post-name").value.trim();
  const msg  = document.getElementById("post-msg").value.trim();
  const btn  = document.getElementById("submit-btn");
  const ok   = document.getElementById("msg-ok");
  const err  = document.getElementById("msg-err");
  ok.style.display = "none"; err.style.display = "none";
  if (!name || !msg) {{ err.textContent = "Please enter your name and message."; err.style.display = "block"; return; }}
  btn.disabled = true; btn.textContent = "Sending...";
  try {{
    const res = await fetch("/api/posts", {{
      method: "POST",
      headers: {{"Content-Type": "application/json"}},
      body: JSON.stringify({{name, message: msg}})
    }});
    const d = await res.json();
    if (!d.ok) throw new Error(d.error || "Failed to submit.");
    ok.style.display = "block";
    document.getElementById("post-name").value = "";
    document.getElementById("post-msg").value = "";
    loadPosts();
  }} catch(e) {{
    err.textContent = "Error: " + (e.message || e);
    err.style.display = "block";
  }} finally {{
    btn.disabled = false; btn.textContent = "&#x1F4E8; Submit";
  }}
}}

async function loadPosts() {{
  try {{
    const res = await fetch("/api/posts");
    const d = await res.json();
    const wrap = document.getElementById("posts-list");
    if (!d.ok || !d.posts.length) {{
      wrap.innerHTML = '<p class="empty-board">&#x1F4ED; No posts yet. Be the first to post!</p>';
      return;
    }}
    wrap.innerHTML = d.posts.map(p => `
      <div class="post-card">
        <div class="post-header">
          <span class="post-name">&#x1F464; ${{esc(p.name)}}</span>
          <span class="post-time">&#x1F552; ${{p.created_at}}</span>
        </div>
        <div class="post-msg">${{esc(p.message)}}</div>
      </div>
    `).join("");
  }} catch(e) {{}}
}}

function esc(s) {{
  return String(s).replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;").replace(/"/g,"&quot;");
}}

loadPosts();
</script>
</body></html>
"""


def build_guide_html() -> str:
    return f"""<!DOCTYPE html>
<html lang="en"><head>
<meta charset="UTF-8" /><meta name="viewport" content="width=device-width, initial-scale=1" />
<title>How to Create an API Key — AlphaFutures</title>
<style>{_SHARED_CSS}
.guide-wrap {{ max-width: 680px; margin: 0 auto; }}
.step {{ background: var(--surface); border: 1px solid var(--border); border-radius: var(--radius); padding: 24px; margin-bottom: 16px; display: flex; gap: 20px; }}
.step-num {{ min-width: 36px; height: 36px; border-radius: 50%; background: var(--primary); color: #fff; font-size: 16px; font-weight: 700; display: flex; align-items: center; justify-content: center; flex-shrink: 0; }}
.step-body h3 {{ font-size: 15px; font-weight: 700; margin-bottom: 8px; }}
.step-body p {{ font-size: 13px; color: var(--muted); line-height: 1.7; }}
.step-body code {{ background: var(--surface-2); padding: 2px 8px; border-radius: 4px; font-size: 12px; color: var(--text); }}
.tag-ok {{ display: inline-block; background: rgba(16,185,129,.15); color: var(--success); padding: 3px 10px; border-radius: 4px; font-size: 12px; font-weight: 700; margin: 2px 4px 2px 0; }}
.tag-no {{ display: inline-block; background: rgba(239,68,68,.15); color: var(--danger); padding: 3px 10px; border-radius: 4px; font-size: 12px; font-weight: 700; margin: 2px 4px 2px 0; }}
.warn-box {{ background: rgba(245,158,11,.1); border: 1px solid rgba(245,158,11,.3); border-radius: 8px; padding: 14px 16px; margin-top: 12px; font-size: 13px; color: var(--warning); line-height: 1.6; }}
.ref-box {{ background: var(--surface); border: 1px solid var(--border); border-radius: var(--radius); padding: 20px; margin-top: 8px; text-align: center; }}
.ref-box a {{ color: var(--primary); font-size: 14px; }}
.cta {{ display: block; width: 100%; padding: 14px; background: var(--primary); color: #fff; border: none; border-radius: 8px; font-size: 15px; font-weight: 700; text-align: center; text-decoration: none; margin-top: 8px; }}
</style></head><body>
<div class="container">
  <div class="header">
    <div>
      <h1>&#x26A1; AlphaFutures</h1>
      <div style="font-size:11px;color:var(--muted);margin-top:4px;">How to create a Binance API Key</div>
    </div>
    <nav class="nav">
      <a href="/">Dashboard</a>
      <a href="/register">Register</a>
    </nav>
  </div>

  <div class="guide-wrap">
    <div class="step">
      <div class="step-num">1</div>
      <div class="step-body">
        <h3>Log in to Binance</h3>
        <p>Go to <strong>binance.com</strong> and log in to your account. We recommend doing this on a computer for convenience.</p>
      </div>
    </div>

    <div class="step">
      <div class="step-num">2</div>
      <div class="step-body">
        <h3>Go to API Management</h3>
        <p>Click your <strong>profile icon</strong> in the top right → select <strong>API Management</strong> from the dropdown menu.</p>
      </div>
    </div>

    <div class="step">
      <div class="step-num">3</div>
      <div class="step-body">
        <h3>Create a New API Key</h3>
        <p>Click <strong>Create API</strong> → choose <strong>System generated</strong> → set a label such as <code>AlphaFutures</code> → click <strong>Next</strong>.</p>
      </div>
    </div>

    <div class="step">
      <div class="step-num">4</div>
      <div class="step-body">
        <h3>Verify Identity (2FA)</h3>
        <p>Enter the OTP sent to your <strong>Email</strong> and the code from your <strong>Google Authenticator</strong> or other 2FA app → click <strong>Confirm</strong>.</p>
      </div>
    </div>

    <div class="step">
      <div class="step-num">5</div>
      <div class="step-body">
        <h3>Set Permissions</h3>
        <p>Enable and disable exactly as shown below:</p>
        <div style="margin-top:10px;">
          <span class="tag-ok">&#10003; Enable Reading</span>
          <span class="tag-ok">&#10003; Enable Futures</span>
          <span class="tag-no">&#10007; Enable Spot &amp; Margin Trading</span>
          <span class="tag-no">&#10007; Enable Withdrawals</span>
          <span class="tag-no">&#10007; Enable Internal Transfer</span>
        </div>
        <div class="warn-box">&#9888; <strong>Never enable Withdrawals</strong> — if enabled, anyone with the key can withdraw funds from your account.</div>
      </div>
    </div>

    <div class="step">
      <div class="step-num">6</div>
      <div class="step-body">
        <h3>Set IP Restriction (Recommended)</h3>
        <p>Select <strong>Restrict access to trusted IPs only</strong> and enter the server IP: <code>45.77.38.167</code><br>This locks the API Key to our server only, adding an extra layer of security.</p>
      </div>
    </div>

    <div class="step">
      <div class="step-num">7</div>
      <div class="step-body">
        <h3>Copy API Key and Secret Key</h3>
        <p>The screen will show your <strong>API Key</strong> and <strong>Secret Key</strong> — copy both immediately.</p>
        <div class="warn-box">&#9888; <strong>The Secret Key is shown only once.</strong> If you close the page you will not be able to view it again and must create a new one.</div>
      </div>
    </div>

    <div class="step">
      <div class="step-num">8</div>
      <div class="step-body">
        <h3>Enter Keys in the System</h3>
        <p>Paste your API Key and Secret Key on the Register page. The system will start trading automatically after Admin activates your account.</p>
        <a href="/register" class="cta">&#x27A1; Go to Register</a>
      </div>
    </div>

    <div class="ref-box">
      <p style="font-size:12px;color:var(--muted);margin-bottom:8px;">Reference: Binance Official Documentation</p>
      <a href="https://www.binance.com/en/support/faq/detail/360002502072" target="_blank">binance.com/en/support/faq/detail/360002502072 &#x2197;</a>
    </div>
  </div>
</div>
</body></html>
"""


def build_login_html() -> str:
    return f"""<!DOCTYPE html>
<html lang="en"><head>
<meta charset="UTF-8" /><meta name="viewport" content="width=device-width, initial-scale=1" />
<title>Login — AlphaFutures</title>
<style>{_SHARED_CSS}
.form-wrap {{ max-width: 400px; margin: 60px auto; }}
.form-card {{ background: var(--surface); border: 1px solid var(--border); border-radius: var(--radius); padding: 32px; }}
.form-card h2 {{ font-size: 20px; font-weight: 700; margin-bottom: 6px; }}
.form-card .sub {{ color: var(--muted); font-size: 13px; margin-bottom: 24px; }}
.field-label {{ display: block; font-size: 11px; color: var(--muted); text-transform: uppercase; letter-spacing: .5px; margin-bottom: 6px; font-weight: 600; }}
.field-input {{ width: 100%; background: var(--bg); border: 1px solid var(--border); border-radius: 8px; padding: 11px 13px; color: var(--text); font-size: 14px; margin-bottom: 14px; outline: none; font-family: inherit; }}
.field-input:focus {{ border-color: var(--primary); }}
.submit-btn {{ width: 100%; padding: 13px; background: var(--primary); color: #fff; border: none; border-radius: 8px; font-size: 15px; font-weight: 700; cursor: pointer; transition: .15s; }}
.submit-btn:hover {{ background: #2563eb; }}
.submit-btn:disabled {{ opacity: .5; cursor: not-allowed; }}
.msg {{ margin-top: 14px; padding: 12px 14px; border-radius: 8px; font-size: 13px; display: none; }}
.msg-err {{ background: rgba(239,68,68,.12); color: var(--danger); }}
.divider {{ text-align: center; color: var(--muted); font-size: 13px; margin: 18px 0 0; }}
.divider a {{ color: var(--primary); text-decoration: none; }}
</style></head><body>
<div class="container">
  <div class="header">
    <div>
      <h1>&#x26A1; AlphaFutures</h1>
      <div style="font-size:11px;color:var(--muted);margin-top:4px;">Member Login</div>
    </div>
    <nav class="nav">
      <a href="/">Dashboard</a>
      <a href="/register">Join</a>
    </nav>
  </div>

  <div class="form-wrap">
    <div class="form-card">
      <h2>&#x1F464; Member Login</h2>
      <div class="sub">Sign in to access your personal dashboard.</div>
      <label class="field-label">Email</label>
      <input class="field-input" type="email" id="email" placeholder="your@email.com" autocomplete="email" onkeydown="if(event.key==='Enter')login()" />
      <label class="field-label">Password</label>
      <input class="field-input" type="password" id="pw" placeholder="Your password" autocomplete="current-password" onkeydown="if(event.key==='Enter')login()" />
      <button class="submit-btn" id="btn" onclick="login()">&#x27A1; Login</button>
      <div class="msg msg-err" id="err"></div>
      <div class="divider">Don't have an account? <a href="/register">Register here</a></div>
    </div>
  </div>
</div>
<script>
async function login() {{
  const email = document.getElementById('email').value.trim();
  const pw    = document.getElementById('pw').value;
  const btn   = document.getElementById('btn');
  const err   = document.getElementById('err');
  err.style.display = 'none';
  if (!email || !pw) {{ showErr('Please enter your email and password.'); return; }}
  btn.disabled = true; btn.textContent = 'Signing in...';
  try {{
    const res = await fetch('/api/login', {{
      method: 'POST',
      headers: {{'Content-Type': 'application/json'}},
      body: JSON.stringify({{email, password: pw}})
    }});
    const d = await res.json();
    if (!d.ok) {{ showErr(d.error || 'Invalid email or password.'); btn.disabled=false; btn.textContent='→ Login'; return; }}
    window.location.href = '/u/' + d.token;
  }} catch(e) {{ showErr('Network error: ' + e.message); btn.disabled=false; btn.textContent='→ Login'; }}
}}
function showErr(msg) {{
  const el = document.getElementById('err');
  el.textContent = msg; el.style.display = 'block';
}}
</script>
</body></html>
"""


_FORM_CSS = """
.form-wrap { max-width: 440px; margin: 32px auto; }
.form-card { background: var(--surface); border: 1px solid var(--border); border-radius: var(--radius); padding: 32px; }
.form-card h2 { font-size: 20px; font-weight: 700; margin-bottom: 6px; }
.form-card .sub { color: var(--muted); font-size: 13px; margin-bottom: 24px; line-height: 1.6; }
.field-label { display: block; font-size: 11px; color: var(--muted); text-transform: uppercase; letter-spacing: .5px; margin-bottom: 6px; font-weight: 600; }
.field-input { width: 100%; background: var(--bg); border: 1px solid var(--border); border-radius: 8px; padding: 11px 13px; color: var(--text); font-size: 14px; margin-bottom: 14px; outline: none; font-family: inherit; }
.field-input:focus { border-color: var(--primary); }
.submit-btn { width: 100%; padding: 13px; background: var(--primary); color: #fff; border: none; border-radius: 8px; font-size: 15px; font-weight: 700; cursor: pointer; transition: .15s; }
.submit-btn:hover { background: #2563eb; }
.submit-btn:disabled { opacity: .5; cursor: not-allowed; }
.msg { margin-top: 14px; padding: 12px 14px; border-radius: 8px; font-size: 13px; display: none; line-height: 1.7; }
.msg-ok { background: rgba(16,185,129,.12); color: var(--success); }
.msg-err { background: rgba(239,68,68,.12); color: var(--danger); }
.step-bar { display: flex; gap: 0; margin-bottom: 28px; }
.step-item { flex: 1; text-align: center; font-size: 12px; font-weight: 600; padding: 8px 4px; border-bottom: 3px solid var(--border); color: var(--muted); }
.step-item.active { border-bottom-color: var(--primary); color: var(--primary); }
.step-item.done { border-bottom-color: var(--success); color: var(--success); }
.price-box { background: rgba(59,130,246,.08); border: 1px solid rgba(59,130,246,.2); border-radius: 8px; padding: 12px 14px; margin-bottom: 20px; font-size: 13px; color: var(--muted); line-height: 1.6; }
.price-box strong { color: var(--primary); font-size: 15px; }
"""


def build_register_html() -> str:
    return f"""<!DOCTYPE html>
<html lang="en"><head>
<meta charset="UTF-8" /><meta name="viewport" content="width=device-width, initial-scale=1" />
<title>Register — AlphaFutures</title>
<style>{_SHARED_CSS}{_FORM_CSS}</style>
</head><body>
<div class="container">
  <div class="header">
    <div>
      <h1>&#x26A1; AlphaFutures</h1>
      <div style="font-size:11px;color:var(--muted);margin-top:3px;">Register — Elliott Wave Auto Trading</div>
    </div>
    <nav class="nav"><a href="/">Dashboard</a></nav>
  </div>

  <div class="form-wrap">
    <div class="step-bar">
      <div class="step-item active" id="step1-tab">1 — Create Account</div>
      <div class="step-item" id="step2-tab">2 — Connect Binance</div>
      <div class="step-item" id="step3-tab">3 — Pending</div>
    </div>

    <!-- Step 1 -->
    <div id="step1" class="form-card">
      <h2>&#x1F464; Create Account</h2>
      <div class="sub">Enter your email and set a password to get started.</div>
      <div class="price-box">&#x1F4B0; Service fee <strong>300 THB / month</strong> — Admin will contact you after registration.</div>
      <label class="field-label">Email</label>
      <input class="field-input" type="email" id="s1-email" placeholder="your@email.com" autocomplete="email" />
      <label class="field-label">Password</label>
      <input class="field-input" type="password" id="s1-pw" placeholder="At least 8 characters" />
      <label class="field-label">Confirm Password</label>
      <input class="field-input" type="password" id="s1-pw2" placeholder="Re-enter your password" />
      <button class="submit-btn" id="s1-btn" onclick="step1()">Next &#x2192;</button>
      <div class="msg msg-err" id="s1-err"></div>
    </div>

    <!-- Step 2 -->
    <div id="step2" class="form-card" style="display:none">
      <h2>&#x1F511; Connect Binance</h2>
      <div class="sub">Enter your Binance Futures API Key and Secret.<br>Don't have one? <a href="/guide" style="color:var(--primary);">See how to create an API Key</a></div>
      <label class="field-label">Binance API Key</label>
      <input class="field-input" type="text" id="s2-key" placeholder="Paste your API Key here" autocomplete="off" spellcheck="false" style="font-family:monospace;font-size:12px;" />
      <label class="field-label">Binance API Secret</label>
      <input class="field-input" type="password" id="s2-secret" placeholder="Paste your Secret Key here" autocomplete="off" style="font-family:monospace;font-size:12px;" />
      <button class="submit-btn" id="s2-btn" onclick="step2()">&#x2713; Submit</button>
      <div class="msg msg-err" id="s2-err"></div>
    </div>

    <!-- Step 3 -->
    <div id="step3" class="form-card" style="display:none;text-align:center;">
      <div style="font-size:48px;margin-bottom:16px;">&#x23F3;</div>
      <h2 style="margin-bottom:12px;">Pending Admin Approval</h2>
      <div class="sub" style="margin-bottom:20px;">Admin will activate your account within 24 hours.<br>After activation you can access your personal Dashboard.</div>
      <div style="background:var(--surface-2);border-radius:8px;padding:14px;margin-bottom:16px;">
        <div style="font-size:11px;color:var(--muted);margin-bottom:6px;">Your Dashboard link — save this</div>
        <a id="dash-link" href="#" style="color:var(--primary);font-size:13px;word-break:break-all;"></a>
      </div>
      <a id="dash-btn" href="#" class="submit-btn" style="display:inline-block;text-decoration:none;padding:12px 24px;width:auto;">&#x1F4CA; Go to Dashboard</a>
    </div>
  </div>
</div>
<script>
let _token = '';

async function step1() {{
  const email = document.getElementById('s1-email').value.trim();
  const pw    = document.getElementById('s1-pw').value;
  const pw2   = document.getElementById('s1-pw2').value;
  const err   = document.getElementById('s1-err');
  const btn   = document.getElementById('s1-btn');
  err.style.display = 'none';
  if (!email || !pw || !pw2) {{ showErr('s1-err','Please fill in all fields.'); return; }}
  if (!email.includes('@')) {{ showErr('s1-err','Invalid email address.'); return; }}
  if (pw.length < 8) {{ showErr('s1-err','Password must be at least 8 characters.'); return; }}
  if (pw !== pw2) {{ showErr('s1-err','Passwords do not match.'); return; }}
  btn.disabled = true; btn.textContent = 'Creating account...';
  try {{
    const res = await fetch('/api/register/step1', {{
      method:'POST', headers:{{'Content-Type':'application/json'}},
      body: JSON.stringify({{email, password: pw}})
    }});
    const d = await res.json();
    if (!d.ok) {{ showErr('s1-err', d.error || 'Registration failed. Please try again.'); btn.disabled=false; btn.textContent='Next →'; return; }}
    _token = d.token;
    document.getElementById('step1').style.display = 'none';
    document.getElementById('step2').style.display = 'block';
    document.getElementById('step1-tab').className = 'step-item done';
    document.getElementById('step2-tab').className = 'step-item active';
  }} catch(e) {{ showErr('s1-err','Network error: '+e.message); btn.disabled=false; btn.textContent='Next →'; }}
}}

async function step2() {{
  const key    = document.getElementById('s2-key').value.trim();
  const secret = document.getElementById('s2-secret').value.trim();
  const btn    = document.getElementById('s2-btn');
  if (!key || !secret) {{ showErr('s2-err','Please enter your API Key and Secret.'); return; }}
  btn.disabled = true; btn.textContent = 'Saving...';
  try {{
    const res = await fetch('/api/register/step2', {{
      method:'POST', headers:{{'Content-Type':'application/json'}},
      body: JSON.stringify({{token: _token, api_key: key, api_secret: secret}})
    }});
    const d = await res.json();
    if (!d.ok) {{ showErr('s2-err', d.error || 'Failed to save. Please try again.'); btn.disabled=false; btn.textContent='✓ Submit'; return; }}
    const url = window.location.origin + '/u/' + _token;
    document.getElementById('dash-link').href = url;
    document.getElementById('dash-link').textContent = url;
    document.getElementById('dash-btn').href = '/u/' + _token;
    document.getElementById('step2').style.display = 'none';
    document.getElementById('step3').style.display = 'block';
    document.getElementById('step2-tab').className = 'step-item done';
    document.getElementById('step3-tab').className = 'step-item active';
  }} catch(e) {{ showErr('s2-err','Network error: '+e.message); btn.disabled=false; btn.textContent='✓ Submit'; }}
}}

function showErr(id, msg) {{
  const el = document.getElementById(id);
  el.textContent = msg; el.style.display = 'block';
}}
</script>
</body></html>
"""


def build_admin_html() -> str:
    return f"""<!DOCTYPE html>
<html lang="en"><head>
<meta charset="UTF-8" /><meta name="viewport" content="width=device-width, initial-scale=1" />
<title>Admin — AlphaFutures</title>
<style>{_SHARED_CSS}
.login-card {{ max-width: 360px; margin: 100px auto; background: var(--surface); border: 1px solid var(--border); border-radius: var(--radius); padding: 32px; }}
.login-card h2 {{ font-size: 18px; font-weight: 700; margin-bottom: 20px; }}
label {{ display: block; font-size: 12px; color: var(--muted); text-transform: uppercase; letter-spacing: .5px; margin-bottom: 6px; }}
input[type=password], input[type=text] {{ width: 100%; background: var(--surface-2); border: 1px solid var(--border); border-radius: 8px; padding: 10px 12px; color: var(--text); font-size: 14px; margin-bottom: 12px; outline: none; font-family: monospace; }}
input:focus {{ border-color: var(--primary); }}
.btn {{ padding: 7px 13px; border: none; border-radius: 7px; font-size: 12px; font-weight: 600; cursor: pointer; white-space: nowrap; }}
.btn-primary {{ background: var(--primary); color: #fff; }}
.btn-success {{ background: var(--success); color: #fff; }}
.btn-danger {{ background: var(--danger); color: #fff; }}
.btn-warn {{ background: var(--warning); color: #000; }}
.btn-ghost {{ background: var(--surface-2); color: var(--muted); }}
.btn:disabled {{ opacity: .5; cursor: not-allowed; }}
.badge-active {{ background: rgba(16,185,129,.15); color: var(--success); padding: 3px 8px; border-radius: 4px; font-size: 11px; font-weight: 600; }}
.badge-inactive {{ background: rgba(156,163,175,.15); color: var(--muted); padding: 3px 8px; border-radius: 4px; font-size: 11px; font-weight: 600; }}
.badge-paid {{ background: rgba(16,185,129,.15); color: var(--success); padding: 3px 8px; border-radius: 4px; font-size: 11px; font-weight: 600; }}
.badge-overdue {{ background: rgba(239,68,68,.15); color: var(--danger); padding: 3px 8px; border-radius: 4px; font-size: 11px; font-weight: 600; }}
.badge-unpaid {{ background: rgba(245,158,11,.12); color: var(--warning); padding: 3px 8px; border-radius: 4px; font-size: 11px; font-weight: 600; }}
.badge-admin {{ background: rgba(59,130,246,.15); color: var(--primary); padding: 3px 8px; border-radius: 4px; font-size: 11px; font-weight: 600; }}
#main {{ display: none; }}
.err {{ color: var(--danger); font-size: 13px; margin-top: 8px; }}
.summary-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(120px,1fr)); gap: 10px; margin-bottom: 16px; }}
.summary-card {{ background: var(--surface); border: 1px solid var(--border); border-radius: 10px; padding: 14px 16px; }}
.summary-card .val {{ font-size: 28px; font-weight: 700; }}
.summary-card .lbl {{ font-size: 11px; color: var(--muted); text-transform: uppercase; letter-spacing: .4px; margin-top: 2px; }}
.api-edit {{ background: var(--surface-2); border-radius: 8px; padding: 12px; margin-top: 8px; display: none; }}
.api-edit input {{ margin-bottom: 8px; font-size: 12px; padding: 8px 10px; }}
.actions {{ display: flex; gap: 5px; flex-wrap: wrap; }}
tr.row-admin {{ background: rgba(59,130,246,.04); }}
</style></head><body>
<div class="container">
  <div id="login-view">
    <div class="login-card">
      <h2>&#x1F512; Admin Login</h2>
      <label>Password</label>
      <input type="password" id="pw" placeholder="Enter admin password" onkeydown="if(event.key==='Enter')login()" style="font-family:inherit;" />
      <button class="btn btn-primary" onclick="login()" style="width:100%;padding:10px;font-size:14px;">Login</button>
      <div class="err" id="login-err"></div>
    </div>
  </div>

  <div id="main">
    <div class="header">
      <h1>&#x26A1; Admin Panel</h1>
      <nav class="nav">
        <a href="/">Dashboard</a>
        <a href="/admin" class="active">Admin</a>
      </nav>
    </div>

    <div class="summary-grid">
      <div class="summary-card"><div class="val" id="s-total">0</div><div class="lbl">Total Members</div></div>
      <div class="summary-card"><div class="val text-success" id="s-active">0</div><div class="lbl">Active</div></div>
      <div class="summary-card"><div class="val text-success" id="s-paid">0</div><div class="lbl">Paid</div></div>
      <div class="summary-card"><div class="val text-danger" id="s-overdue">0</div><div class="lbl">Overdue</div></div>
      <div class="summary-card"><div class="val text-warning" id="s-unpaid">0</div><div class="lbl">Never Paid</div></div>
    </div>

    <div class="section">
      <div class="section-head"><h2>&#x1F465; Members</h2></div>
      <div class="table-wrap">
        <table class="table">
          <thead><tr>
            <th>Name</th><th>Email</th><th>Status</th><th>Payment</th>
            <th>Paid Until</th><th>Days Left</th><th>API Key</th><th>Joined</th><th>Actions</th>
          </tr></thead>
          <tbody id="accounts-tbody"></tbody>
        </table>
      </div>
    </div>
  </div>
</div>

<script>
let _pw = '';
let _accounts = [];

async function login() {{
  _pw = document.getElementById('pw').value;
  const res = await fetch('/api/admin/accounts', {{ headers: {{'X-Admin-Password': _pw}} }});
  if (res.status === 401) {{ document.getElementById('login-err').textContent = 'Wrong password'; return; }}
  const d = await res.json();
  if (!d.ok) {{ document.getElementById('login-err').textContent = d.error || 'Error'; return; }}
  document.getElementById('login-view').style.display = 'none';
  document.getElementById('main').style.display = 'block';
  renderAccounts(d.accounts);
}}

function renderAccounts(accounts) {{
  _accounts = accounts;
  const members = accounts.filter(a => a.role !== 'admin');
  document.getElementById('s-total').textContent = members.length;
  document.getElementById('s-active').textContent = members.filter(a=>a.active).length;
  document.getElementById('s-paid').textContent = members.filter(a=>a.payment_status==='paid').length;
  document.getElementById('s-overdue').textContent = members.filter(a=>a.payment_status==='overdue').length;
  document.getElementById('s-unpaid').textContent = members.filter(a=>a.payment_status==='unpaid').length;

  const tbody = document.getElementById('accounts-tbody');
  if (!accounts.length) {{
    tbody.innerHTML = '<tr><td colspan="9" style="text-align:center;color:var(--muted);padding:32px">No members yet</td></tr>';
    return;
  }}

  tbody.innerHTML = accounts.map(a => {{
    const isAdmin = a.role === 'admin';
    const payBadge = isAdmin
      ? `<span class="badge-admin">ADMIN</span>`
      : a.payment_status === 'paid'
        ? `<span class="badge-paid">Paid</span>`
        : a.payment_status === 'overdue'
          ? `<span class="badge-overdue">Overdue</span>`
          : `<span class="badge-unpaid">Unpaid</span>`;

    const daysLeft = a.days_remaining > 0
      ? `<span class="text-success">${{a.days_remaining}} วัน</span>`
      : a.paid_until
        ? `<span class="text-danger">${{a.days_remaining}} วัน</span>`
        : '—';

    const activateBtn = isAdmin ? '' : a.active
      ? `<button class="btn btn-danger" onclick="toggleActive(${{a.id}},false)">Deactivate</button>`
      : `<button class="btn btn-success" onclick="toggleActive(${{a.id}},true)">Activate</button>`;

    const paidBtn = isAdmin ? '' :
      `<button class="btn btn-warn" onclick="markPaid(${{a.id}})">+1 Month</button>`;

    const apiBtn = `<button class="btn btn-ghost" onclick="toggleApiEdit(${{a.id}})">&#x270F; API</button>`;

    return `
      <tr class="${{isAdmin ? 'row-admin' : ''}}" id="row-${{a.id}}">
        <td style="font-weight:600">${{esc(a.label)}}</td>
        <td style="font-size:12px;color:var(--muted)">${{esc(a.email)}}</td>
        <td><span class="${{a.active ? 'badge-active' : 'badge-inactive'}}">${{a.active ? 'ACTIVE' : 'INACTIVE'}}</span></td>
        <td>${{payBadge}}</td>
        <td style="font-size:12px;color:var(--muted)">${{a.paid_until || '—'}}</td>
        <td style="font-size:12px">${{daysLeft}}</td>
        <td style="font-size:11px;font-family:monospace;color:var(--muted)">${{a.has_api_key ? a.api_key_masked : '<span style="color:var(--danger)">None</span>'}}</td>
        <td style="font-size:11px;color:var(--muted)">${{(a.created_at||'').substring(0,10)}}</td>
        <td>
          <div class="actions">
            ${{activateBtn}}
            ${{paidBtn}}
            ${{apiBtn}}
          </div>
          <div class="api-edit" id="api-edit-${{a.id}}">
            <input type="text" id="akey-${{a.id}}" placeholder="API Key" />
            <input type="text" id="asec-${{a.id}}" placeholder="API Secret" />
            <div style="display:flex;gap:6px;">
              <button class="btn btn-primary" onclick="saveApi(${{a.id}})">Save</button>
              <button class="btn btn-ghost" onclick="toggleApiEdit(${{a.id}})">Cancel</button>
            </div>
          </div>
        </td>
      </tr>`;
  }}).join('');
}}

function esc(s) {{
  return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
}}

function toggleApiEdit(id) {{
  const el = document.getElementById('api-edit-'+id);
  el.style.display = el.style.display === 'none' || !el.style.display ? 'block' : 'none';
}}

async function reload() {{
  const res = await fetch('/api/admin/accounts', {{ headers: {{'X-Admin-Password': _pw}} }});
  const d = await res.json();
  if (d.ok) renderAccounts(d.accounts);
}}

async function toggleActive(id, activate) {{
  const ep = activate ? '/api/admin/activate' : '/api/admin/deactivate';
  await fetch(ep, {{ method:'POST', headers:{{'Content-Type':'application/json','X-Admin-Password':_pw}}, body:JSON.stringify({{id}}) }});
  reload();
}}

async function markPaid(id) {{
  const res = await fetch('/api/admin/mark_paid', {{
    method: 'POST',
    headers: {{'Content-Type':'application/json','X-Admin-Password':_pw}},
    body: JSON.stringify({{id}})
  }});
  const d = await res.json();
  if (!d.ok) alert(d.error || 'Error');
  reload();
}}

async function saveApi(id) {{
  const api_key = document.getElementById('akey-'+id).value.trim();
  const api_secret = document.getElementById('asec-'+id).value.trim();
  if (!api_key || !api_secret) {{ alert('Please enter both API Key and Secret.'); return; }}
  const res = await fetch('/api/admin/update_api', {{
    method: 'POST',
    headers: {{'Content-Type':'application/json','X-Admin-Password':_pw}},
    body: JSON.stringify({{id, api_key, api_secret}})
  }});
  const d = await res.json();
  if (d.ok) {{ alert('API Key saved successfully.'); reload(); }}
  else alert(d.error || 'Error');
}}
</script>
</body></html>
"""


def build_fund_html() -> str:
    return f"""<!DOCTYPE html>
<html lang="en"><head>
<meta charset="UTF-8"/><meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>Fund Management — AlphaFutures</title>
<style>{_SHARED_CSS}
.fund-tabs{{display:flex;gap:0;margin-bottom:16px;border-bottom:2px solid var(--border);}}
.tab{{padding:10px 20px;cursor:pointer;font-size:13px;font-weight:600;color:var(--muted);border-bottom:2px solid transparent;margin-bottom:-2px;}}
.tab.active{{color:var(--primary);border-bottom-color:var(--primary);}}
.tab-panel{{display:none;}}.tab-panel.active{{display:block;}}
.login-card{{max-width:360px;margin:80px auto;background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);padding:32px;}}
.login-card h2{{font-size:18px;font-weight:700;margin-bottom:20px;}}
label{{display:block;font-size:12px;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;margin-bottom:6px;}}
input[type=password],input[type=text],input[type=date],input[type=number],select{{width:100%;background:var(--surface-2);border:1px solid var(--border);border-radius:8px;padding:9px 12px;color:var(--text);font-size:13px;margin-bottom:10px;outline:none;font-family:inherit;}}
input:focus,select:focus{{border-color:var(--primary);}}
.btn{{padding:7px 13px;border:none;border-radius:7px;font-size:12px;font-weight:600;cursor:pointer;white-space:nowrap;}}
.btn-primary{{background:var(--primary);color:#fff;}}.btn-success{{background:var(--success);color:#fff;}}
.btn-danger{{background:var(--danger);color:#fff;}}.btn-ghost{{background:var(--surface-2);color:var(--muted);}}
.btn-warn{{background:var(--warning);color:#000;}}
.err{{color:var(--danger);font-size:13px;margin-top:6px;}}
.sheet-wrap{{overflow-x:auto;}}
.sheet{{border-collapse:collapse;font-size:12px;width:100%;min-width:900px;}}
.sheet th{{background:var(--surface-2);color:var(--muted);font-size:11px;text-transform:uppercase;letter-spacing:.4px;padding:8px 12px;border:1px solid var(--border);text-align:center;position:sticky;top:0;z-index:1;}}
.sheet td{{padding:7px 12px;border:1px solid var(--border);text-align:right;}}
.sheet td:first-child,.sheet td:nth-child(2){{text-align:left;}}
.sheet tr:hover td{{background:rgba(59,130,246,.06);}}
.sheet tfoot td{{font-weight:700;background:var(--surface-2);border-top:2px solid var(--border);}}
.pos{{color:var(--success);font-weight:600;}}.neg{{color:var(--danger);font-weight:600;}}
.badge-long{{background:rgba(16,185,129,.15);color:var(--success);padding:2px 7px;border-radius:4px;font-size:10px;font-weight:700;}}
.badge-short{{background:rgba(239,68,68,.15);color:var(--danger);padding:2px 7px;border-radius:4px;font-size:10px;font-weight:700;}}
.form-row{{display:grid;grid-template-columns:1fr 1fr;gap:10px;}}
.export-bar{{display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;flex-wrap:wrap;gap:8px;}}
#main{{display:none;}}
</style></head><body>
<div class="container">

<div id="login-view">
  <div class="login-card">
    <h2>&#x1F4B0; Fund Admin</h2>
    <label>Password</label>
    <input type="password" id="pw" placeholder="Admin password" onkeydown="if(event.key==='Enter')login()"/>
    <button class="btn btn-primary" onclick="login()" style="width:100%;padding:10px;font-size:14px;">Login</button>
    <div class="err" id="login-err"></div>
  </div>
</div>

<div id="main">
  <div class="header">
    <div><h1>&#x1F4B0; Fund Management</h1>
    <div style="font-size:11px;color:var(--muted);margin-top:4px;">Pooled Fund — AlphaFutures</div></div>
    <nav class="nav"><a href="/">Dashboard</a><a href="/admin">Admin</a></nav>
  </div>

  <div class="fund-tabs">
    <div class="tab active" onclick="showTab('members')">&#x1F465; Members</div>
    <div class="tab" onclick="showTab('trades')">&#x1F4CA; Trades</div>
    <div class="tab" onclick="showTab('settlement')">&#x1F4CB; Settlement Sheet</div>
  </div>

  <!-- MEMBERS TAB -->
  <div id="tab-members" class="tab-panel active">
    <div class="section">
      <div class="section-head">
        <h2>Fund Members</h2>
        <button class="btn btn-success" onclick="toggleForm('add-member-form')">+ Add Member</button>
      </div>
      <div id="add-member-form" style="display:none;background:var(--surface-2);border-radius:10px;padding:16px;margin-bottom:16px;">
        <div class="form-row">
          <div><label>Name</label><input type="text" id="f-name" placeholder="Full name"/></div>
          <div><label>Phone</label><input type="text" id="f-phone" placeholder="Phone number"/></div>
        </div>
        <div class="form-row">
          <div><label>Email</label><input type="text" id="f-email" placeholder="email@example.com"/></div>
          <div><label>Deposit (USDT)</label><input type="number" id="f-deposit" placeholder="e.g. 1000" min="100"/></div>
        </div>
        <div class="form-row">
          <div><label>Join Date</label><input type="date" id="f-joined"/></div>
          <div><label>Note</label><input type="text" id="f-note" placeholder="Optional"/></div>
        </div>
        <button class="btn btn-primary" onclick="addMember()">Save Member</button>
        <span class="err" id="add-err"></span>
      </div>
      <div class="sheet-wrap">
        <table class="sheet">
          <thead><tr><th>#</th><th>Name</th><th>Email</th><th>Phone</th><th>Deposit</th><th>Balance</th><th>P&amp;L</th><th>Return%</th><th>Joined</th><th>Status</th><th>Actions</th></tr></thead>
          <tbody id="members-tbody"><tr><td colspan="11" style="text-align:center;padding:24px;color:var(--muted)">Loading...</td></tr></tbody>
        </table>
      </div>
    </div>
  </div>

  <!-- TRADES TAB -->
  <div id="tab-trades" class="tab-panel">
    <div class="section">
      <div class="section-head">
        <h2>Trades</h2>
        <button class="btn btn-success" onclick="toggleForm('add-trade-form')">+ Add Trade</button>
      </div>
      <div id="add-trade-form" style="display:none;background:var(--surface-2);border-radius:10px;padding:16px;margin-bottom:16px;">
        <div class="form-row">
          <div><label>Symbol</label><input type="text" id="t-symbol" placeholder="e.g. ETHUSDT"/></div>
          <div><label>Timeframe</label><input type="text" id="t-tf" placeholder="4H / 1D"/></div>
        </div>
        <div class="form-row">
          <div><label>Side</label><select id="t-side"><option>LONG</option><option>SHORT</option></select></div>
          <div><label>Open Date</label><input type="text" id="t-opened" placeholder="2026-03-18 06:00:00"/></div>
        </div>
        <div class="form-row">
          <div><label>Entry Price</label><input type="number" id="t-entry" step="any"/></div>
          <div><label>SL Price</label><input type="number" id="t-sl" step="any"/></div>
        </div>
        <button class="btn btn-primary" onclick="addTrade()">Open Trade</button>
        <span class="err" id="trade-add-err"></span>
      </div>
      <div class="sheet-wrap">
        <table class="sheet">
          <thead><tr><th>#</th><th>Symbol</th><th>TF</th><th>Side</th><th>Entry</th><th>SL</th><th>Risk%</th><th>Opened</th><th>Closed</th><th>Result</th><th>RR</th><th>PnL%</th><th>Actions</th></tr></thead>
          <tbody id="trades-tbody"><tr><td colspan="13" style="text-align:center;padding:24px;color:var(--muted)">Loading...</td></tr></tbody>
        </table>
      </div>
    </div>
  </div>

  <!-- SETTLEMENT TAB -->
  <div id="tab-settlement" class="tab-panel">
    <div class="section">
      <div class="export-bar">
        <div style="display:flex;align-items:center;gap:10px;">
          <h2>&#x1F4CB; Settlement Sheet</h2>
          <select id="month-select" onchange="loadSettlement()" style="width:auto;margin:0;font-size:13px;padding:6px 10px;"></select>
        </div>
        <button class="btn btn-success" onclick="exportExcel()">&#x2B07; Export Excel</button>
      </div>

      <div class="sheet-wrap">
        <table class="sheet" id="settlement-table">
          <thead><tr>
            <th>#</th><th>Name</th><th>Deposit</th><th>Balance</th>
            <th>Trades</th><th>Wins</th><th>Losses</th>
            <th>Gross P&amp;L</th><th>Fee 3%</th><th>Net P&amp;L</th><th>Return%</th>
          </tr></thead>
          <tbody id="settlement-tbody"><tr><td colspan="11" style="text-align:center;padding:32px;color:var(--muted)">Select a month above</td></tr></tbody>
          <tfoot id="settlement-tfoot"></tfoot>
        </table>
      </div>

      <div id="detail-section" style="display:none;margin-top:16px;">
        <div class="section-head">
          <h2 id="detail-title">Trade Detail</h2>
          <button class="btn btn-ghost" onclick="closeDetail()">✕ Close</button>
        </div>
        <div class="sheet-wrap">
          <table class="sheet" id="detail-table">
            <thead><tr><th>#</th><th>Symbol</th><th>TF</th><th>Side</th><th>Opened</th><th>Closed</th><th>Result</th><th>RR</th><th>PnL%</th><th>Balance Before</th><th>P&amp;L ($)</th><th>Balance After</th></tr></thead>
            <tbody id="detail-tbody"></tbody>
          </table>
        </div>
      </div>
    </div>
  </div>
</div><!-- #main -->
</div>

<!-- Close Trade Modal -->
<div id="close-modal" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.6);z-index:100;align-items:center;justify-content:center;">
  <div style="background:var(--surface);border-radius:12px;padding:24px;width:340px;max-width:90vw;">
    <h3 style="margin-bottom:16px;">Close Trade</h3>
    <input type="hidden" id="close-trade-id"/>
    <label>Close Date</label><input type="text" id="close-date" placeholder="2026-04-01 12:00:00"/>
    <label>Result</label>
    <select id="close-result">
      <option>TP3_HIT</option><option>TP2_THEN_SL</option><option>TP1_THEN_SL</option><option>SL_HIT</option>
    </select>
    <label>Realized RR</label><input type="number" id="close-rr" step="any" placeholder="e.g. 1.08 or -1"/>
    <div style="display:flex;gap:8px;margin-top:8px;">
      <button class="btn btn-primary" onclick="closeTrade()">Confirm</button>
      <button class="btn btn-ghost" onclick="document.getElementById('close-modal').style.display='none'">Cancel</button>
    </div>
    <div class="err" id="close-err"></div>
  </div>
</div>

<script>
let _pw = '';
let _members = [];
let _trades = [];
let _settlementData = [];
let _currentMonth = '';

async function login() {{
  _pw = document.getElementById('pw').value;
  const res = await fetch('/api/admin/accounts', {{headers:{{'X-Admin-Password':_pw}}}});
  if (res.status === 401) {{ document.getElementById('login-err').textContent='Wrong password'; return; }}
  document.getElementById('login-view').style.display = 'none';
  document.getElementById('main').style.display = 'block';
  await Promise.all([loadMembers(), loadTrades(), loadMonths()]);
}}

function showTab(name) {{
  document.querySelectorAll('.tab').forEach((t,i) => t.classList.toggle('active', ['members','trades','settlement'][i]===name));
  document.querySelectorAll('.tab-panel').forEach(p => p.classList.remove('active'));
  document.getElementById('tab-'+name).classList.add('active');
}}

function toggleForm(id) {{
  const el = document.getElementById(id);
  el.style.display = el.style.display==='none' ? 'block' : 'none';
}}

function fmt(v, d=2) {{ return parseFloat(v||0).toFixed(d); }}
function fmtPnl(v) {{
  const n = parseFloat(v||0);
  return `<span class="${{n>=0?'pos':'neg'}}">${{n>=0?'+':''}}${{n.toFixed(2)}}</span>`;
}}
function esc(s) {{ return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;'); }}

// ── Members ──────────────────────────────────────────────────────
async function loadMembers() {{
  const res = await fetch('/api/fund/members', {{headers:{{'X-Admin-Password':_pw}}}});
  const d = await res.json();
  _members = d.members || [];
  const tbody = document.getElementById('members-tbody');
  if (!_members.length) {{ tbody.innerHTML='<tr><td colspan="11" style="text-align:center;padding:24px;color:var(--muted)">No members yet</td></tr>'; return; }}
  tbody.innerHTML = _members.map((m,i) => {{
    const pnl = m.balance_usdt - m.initial_deposit;
    const ret = m.initial_deposit ? (pnl/m.initial_deposit*100) : 0;
    return `<tr>
      <td>${{i+1}}</td>
      <td style="font-weight:600">${{esc(m.name)}}</td>
      <td style="color:var(--muted);font-size:11px">${{esc(m.email)}}</td>
      <td style="color:var(--muted);font-size:11px">${{esc(m.phone)}}</td>
      <td>${{fmt(m.initial_deposit)}}</td>
      <td style="font-weight:600">${{fmt(m.balance_usdt)}}</td>
      <td>${{fmtPnl(pnl)}}</td>
      <td class="${{ret>=0?'pos':'neg'}}">${{ret>=0?'+':''}}${{ret.toFixed(2)}}%</td>
      <td style="color:var(--muted);font-size:11px">${{m.joined_at.substring(0,10)}}</td>
      <td><span style="font-size:11px;font-weight:700;color:${{m.active?'var(--success)':'var(--muted)'}}">${{m.active?'ACTIVE':'INACTIVE'}}</span></td>
      <td><button class="btn btn-danger" onclick="deactivateMember(${{m.id}})" style="font-size:11px">Deactivate</button></td>
    </tr>`;
  }}).join('');
}}

async function addMember() {{
  const body = {{
    name: document.getElementById('f-name').value.trim(),
    email: document.getElementById('f-email').value.trim(),
    phone: document.getElementById('f-phone').value.trim(),
    deposit_usdt: parseFloat(document.getElementById('f-deposit').value),
    joined_at: document.getElementById('f-joined').value,
    note: document.getElementById('f-note').value.trim(),
  }};
  if (!body.name || !body.deposit_usdt || !body.joined_at) {{ document.getElementById('add-err').textContent='Name, deposit and date required'; return; }}
  if (body.deposit_usdt < 100) {{ document.getElementById('add-err').textContent='Minimum deposit is $100'; return; }}
  const res = await fetch('/api/fund/members', {{method:'POST',headers:{{'Content-Type':'application/json','X-Admin-Password':_pw}},body:JSON.stringify(body)}});
  const d = await res.json();
  if (!d.ok) {{ document.getElementById('add-err').textContent = d.error||'Error'; return; }}
  document.getElementById('add-err').textContent='';
  document.getElementById('add-member-form').style.display='none';
  loadMembers();
}}

async function deactivateMember(id) {{
  if (!confirm('Deactivate this member?')) return;
  await fetch('/api/fund/members/deactivate', {{method:'POST',headers:{{'Content-Type':'application/json','X-Admin-Password':_pw}},body:JSON.stringify({{id}})}});
  loadMembers();
}}

// ── Trades ───────────────────────────────────────────────────────
async function loadTrades() {{
  const res = await fetch('/api/fund/trades', {{headers:{{'X-Admin-Password':_pw}}}});
  const d = await res.json();
  _trades = d.trades || [];
  const tbody = document.getElementById('trades-tbody');
  if (!_trades.length) {{ tbody.innerHTML='<tr><td colspan="13" style="text-align:center;padding:24px;color:var(--muted)">No trades yet</td></tr>'; return; }}
  tbody.innerHTML = _trades.map((t,i) => {{
    const isOpen = !t.closed_at;
    const pnlCell = isOpen ? '<td style="color:var(--muted)">OPEN</td>' : `<td class="${{(t.pnl_pct||0)>=0?'pos':'neg'}}">${{((t.pnl_pct||0)*100).toFixed(3)}}%</td>`;
    return `<tr>
      <td>${{i+1}}</td>
      <td style="font-weight:600">${{t.symbol}}</td>
      <td style="color:var(--muted)">${{t.timeframe}}</td>
      <td><span class="badge-${{t.side.toLowerCase()}}">${{t.side}}</span></td>
      <td>${{fmt(t.entry,4)}}</td>
      <td>${{fmt(t.sl,4)}}</td>
      <td style="color:var(--warning)">${{((t.risk_pct||0)*100).toFixed(3)}}%</td>
      <td style="color:var(--muted);font-size:11px">${{t.opened_at.substring(0,16)}}</td>
      <td style="color:var(--muted);font-size:11px">${{t.closed_at?t.closed_at.substring(0,16):'—'}}</td>
      <td style="font-size:11px">${{t.result||'ACTIVE'}}</td>
      <td class="${{(t.realized_rr||0)>=0?'pos':'neg'}}">${{t.realized_rr!=null?fmt(t.realized_rr,4):'—'}}</td>
      ${{pnlCell}}
      <td>${{isOpen?`<button class="btn btn-warn" onclick="openCloseModal(${{t.id}})">Close</button>`:'—'}}</td>
    </tr>`;
  }}).join('');
}}

async function addTrade() {{
  const body = {{
    symbol: document.getElementById('t-symbol').value.trim().toUpperCase(),
    timeframe: document.getElementById('t-tf').value.trim(),
    side: document.getElementById('t-side').value,
    entry: parseFloat(document.getElementById('t-entry').value),
    sl: parseFloat(document.getElementById('t-sl').value),
    opened_at: document.getElementById('t-opened').value.trim(),
  }};
  if (!body.symbol||!body.entry||!body.sl||!body.opened_at) {{ document.getElementById('trade-add-err').textContent='All fields required'; return; }}
  const res = await fetch('/api/fund/trades', {{method:'POST',headers:{{'Content-Type':'application/json','X-Admin-Password':_pw}},body:JSON.stringify(body)}});
  const d = await res.json();
  if (!d.ok) {{ document.getElementById('trade-add-err').textContent=d.error||'Error'; return; }}
  document.getElementById('add-trade-form').style.display='none';
  loadTrades(); loadMonths();
}}

function openCloseModal(id) {{
  document.getElementById('close-trade-id').value = id;
  document.getElementById('close-err').textContent = '';
  document.getElementById('close-modal').style.display = 'flex';
}}

async function closeTrade() {{
  const id = parseInt(document.getElementById('close-trade-id').value);
  const body = {{
    id,
    closed_at: document.getElementById('close-date').value.trim(),
    result: document.getElementById('close-result').value,
    realized_rr: parseFloat(document.getElementById('close-rr').value),
  }};
  if (!body.closed_at || isNaN(body.realized_rr)) {{ document.getElementById('close-err').textContent='Fill all fields'; return; }}
  const res = await fetch('/api/fund/trades/close', {{method:'POST',headers:{{'Content-Type':'application/json','X-Admin-Password':_pw}},body:JSON.stringify(body)}});
  const d = await res.json();
  if (!d.ok) {{ document.getElementById('close-err').textContent=d.error||'Error'; return; }}
  document.getElementById('close-modal').style.display='none';
  loadTrades(); loadMonths(); loadMembers();
}}

// ── Settlement ───────────────────────────────────────────────────
async function loadMonths() {{
  const res = await fetch('/api/fund/months', {{headers:{{'X-Admin-Password':_pw}}}});
  const d = await res.json();
  const sel = document.getElementById('month-select');
  sel.innerHTML = '<option value="">— Select Month —</option>' +
    (d.months||[]).map(m => `<option value="${{m}}">${{m}}</option>`).join('');
  if (d.months && d.months.length) {{ sel.value = d.months[0]; loadSettlement(); }}
}}

async function loadSettlement() {{
  const month = document.getElementById('month-select').value;
  if (!month) return;
  _currentMonth = month;
  const res = await fetch('/api/fund/settlement?month='+month, {{headers:{{'X-Admin-Password':_pw}}}});
  const d = await res.json();
  _settlementData = d.data || [];
  renderSettlement();
}}

function renderSettlement() {{
  const tbody = document.getElementById('settlement-tbody');
  const tfoot = document.getElementById('settlement-tfoot');
  if (!_settlementData.length) {{
    tbody.innerHTML = '<tr><td colspan="11" style="text-align:center;padding:32px;color:var(--muted)">No closed trades for this month</td></tr>';
    tfoot.innerHTML = ''; return;
  }}
  tbody.innerHTML = _settlementData.map((r,i) => `
    <tr style="cursor:pointer" onclick="loadDetail(${{r.id}},'${{esc(r.name)}}')">
      <td>${{i+1}}</td>
      <td style="font-weight:600;color:var(--primary)">${{esc(r.name)}}</td>
      <td>${{fmt(r.initial_deposit)}}</td>
      <td style="font-weight:600">${{fmt(r.balance_end)}}</td>
      <td style="text-align:center">${{r.trades_count}}</td>
      <td style="text-align:center;color:var(--success)">${{r.wins}}</td>
      <td style="text-align:center;color:var(--danger)">${{r.losses}}</td>
      <td>${{fmtPnl(r.gross_pnl)}}</td>
      <td style="color:var(--danger)">-${{fmt(r.fee)}}</td>
      <td style="font-weight:700">${{fmtPnl(r.net_pnl)}}</td>
      <td class="${{r.return_pct>=0?'pos':'neg'}}">${{r.return_pct>=0?'+':''}}${{r.return_pct}}%</td>
    </tr>`).join('');

  const totalGross = _settlementData.reduce((s,r)=>s+r.gross_pnl,0);
  const totalFee   = _settlementData.reduce((s,r)=>s+r.fee,0);
  const totalNet   = _settlementData.reduce((s,r)=>s+r.net_pnl,0);
  tfoot.innerHTML = `<tr>
    <td colspan="7" style="text-align:left">Total</td>
    <td>${{fmtPnl(totalGross)}}</td>
    <td style="color:var(--danger)">-${{fmt(totalFee)}}</td>
    <td style="font-weight:700">${{fmtPnl(totalNet)}}</td>
    <td></td>
  </tr>`;
}}

async function loadDetail(memberId, name) {{
  const month = _currentMonth;
  const res = await fetch(`/api/fund/detail?member_id=${{memberId}}&month=${{month}}`, {{headers:{{'X-Admin-Password':_pw}}}});
  const d = await res.json();
  document.getElementById('detail-title').textContent = name + ' — Trade Detail';
  document.getElementById('detail-section').style.display = 'block';
  const rows = d.trades || [];
  document.getElementById('detail-tbody').innerHTML = rows.map((t,i) => `
    <tr>
      <td>${{i+1}}</td>
      <td style="font-weight:600">${{t.symbol}}</td>
      <td style="color:var(--muted)">${{t.timeframe}}</td>
      <td><span class="badge-${{t.side.toLowerCase()}}">${{t.side}}</span></td>
      <td style="color:var(--muted);font-size:11px">${{t.opened_at.substring(0,16)}}</td>
      <td style="color:var(--muted);font-size:11px">${{t.closed_at.substring(0,16)}}</td>
      <td style="font-size:11px">${{t.result}}</td>
      <td class="${{t.realized_rr>=0?'pos':'neg'}}">${{fmt(t.realized_rr,4)}}</td>
      <td class="${{t.pnl_pct>=0?'pos':'neg'}}">${{(t.pnl_pct*100).toFixed(3)}}%</td>
      <td>${{fmt(t.balance_at_open)}}</td>
      <td>${{fmtPnl(t.pnl_usdt)}}</td>
      <td style="font-weight:600">${{fmt(t.balance_after)}}</td>
    </tr>`).join('');
  document.getElementById('detail-section').scrollIntoView({{behavior:'smooth'}});
}}

function closeDetail() {{
  document.getElementById('detail-section').style.display = 'none';
}}

function exportExcel() {{
  if (!_currentMonth) return;
  window.location.href = '/fund/export/' + _currentMonth + '?pw=' + encodeURIComponent(_pw);
}}
</script>
</body></html>
"""


def run_web_dashboard(
    symbol: str = "BTCUSDT",
    *,
    host: str = "127.0.0.1",
    port: int = 8080,
    refresh_seconds: float = 5.0,
) -> None:

    spa_html = build_spa_html()
    fund_html = build_fund_html()
    db_path = os.getenv("WAVE_DB_PATH", "storage/wave_engine.db")
    account_store = AccountStore(db_path=db_path)
    account_store.seed_admin("jirayuwammagul@gmail.com")
    fund_store = FundStore(db_path=db_path)

    _ensure_posts_table(db_path)
    board_html = build_board_html()

    _visits_path = os.path.join(os.path.dirname(db_path), "visits.json")
    try:
        import json as _j
        with open(_visits_path) as _f:
            _visit_counter: dict = _j.load(_f)
    except Exception:
        _visit_counter = {"count": 0}

    def _save_visits() -> None:
        try:
            import json as _j
            with open(_visits_path, "w") as _f:
                _j.dump(_visit_counter, _f)
        except Exception:
            pass

    _cache: dict = {"payload": None, "updated_at": None}
    _cache_lock = threading.Lock()

    def _refresh_cache() -> None:
        while True:
            try:
                snapshot = build_dashboard_snapshot(symbol)
                snapshot["stats"] = _build_trade_stats(db_path)
                snapshot["active_trades"] = _build_active_trades(db_path)
                updated_at = datetime.now(UTC).strftime("%Y-%m-%d %H:%M:%S")
                payload = {
                    "ok": True,
                    "snapshot": snapshot,
                    "terminal": render_terminal_dashboard(snapshot),
                    "updated_at": updated_at,
                }
            except Exception as exc:
                payload = {"ok": False, "error": str(exc)}
            with _cache_lock:
                _cache["payload"] = payload
                _cache["updated_at"] = datetime.now(UTC)
            import time as _time
            _time.sleep(30)

    t = threading.Thread(target=_refresh_cache, daemon=True)
    t.start()

    class DashboardHandler(BaseHTTPRequestHandler):
        def do_GET(self):  # noqa: N802
            path = self.path.split("?", 1)[0]
            if path in ("/", "/history", "/login", "/register", "/admin", "/guide", "/board"):
                if path == "/":
                    _visit_counter["count"] = _visit_counter.get("count", 0) + 1
                    _save_visits()
                self._send_html(spa_html)
                return
            if path == "/fund":
                self._send_html(fund_html)
                return
            if path.startswith("/fund/export/"):
                month = path[len("/fund/export/"):]
                pw = self.path.split("pw=")[-1] if "pw=" in self.path else ""
                if pw != _ADMIN_PASSWORD:
                    self._send_json(401, {"ok": False, "error": "unauthorized"})
                    return
                self._send_fund_excel(fund_store, month)
                return
            if path == "/api/fund/members":
                if not self._check_admin():
                    return
                members = fund_store.list_members()
                self._send_json(200, {"ok": True, "members": [
                    {"id": m.id, "name": m.name, "email": m.email, "phone": m.phone,
                     "initial_deposit": m.initial_deposit, "balance_usdt": m.balance_usdt,
                     "joined_at": m.joined_at, "active": m.active, "token": m.token}
                    for m in members]})
                return
            if path == "/api/fund/trades":
                if not self._check_admin():
                    return
                trades = fund_store.list_trades()
                self._send_json(200, {"ok": True, "trades": [
                    {"id": t.id, "symbol": t.symbol, "timeframe": t.timeframe, "side": t.side,
                     "entry": t.entry, "sl": t.sl, "opened_at": t.opened_at, "closed_at": t.closed_at,
                     "result": t.result, "realized_rr": t.realized_rr, "risk_pct": t.risk_pct,
                     "pnl_pct": t.pnl_pct, "settlement_month": t.settlement_month}
                    for t in trades]})
                return
            if path == "/api/fund/months":
                if not self._check_admin():
                    return
                self._send_json(200, {"ok": True, "months": fund_store.available_months()})
                return
            if path.startswith("/api/fund/settlement"):
                if not self._check_admin():
                    return
                qs = self.path.split("?", 1)[-1] if "?" in self.path else ""
                month = dict(p.split("=", 1) for p in qs.split("&") if "=" in p).get("month", "")
                data = fund_store.get_settlement_data(month) if month else []
                self._send_json(200, {"ok": True, "data": data})
                return
            if path.startswith("/api/fund/detail"):
                if not self._check_admin():
                    return
                qs = self.path.split("?", 1)[-1] if "?" in self.path else ""
                params = dict(p.split("=", 1) for p in qs.split("&") if "=" in p)
                member_id = int(params.get("member_id", 0))
                month = params.get("month", "")
                trades = fund_store.get_participation_detail(member_id, month)
                self._send_json(200, {"ok": True, "trades": trades})
                return
            if path.startswith("/u/"):
                self._send_html(spa_html)
                return
            if path == "/api/snapshot":
                with _cache_lock:
                    payload = _cache["payload"]
                if payload is None:
                    self._send_json(503, {"ok": False, "error": "snapshot not ready yet, please retry"})
                else:
                    self._send_json(200, payload)
                return
            if path == "/api/history":
                try:
                    trades = _build_trade_history(db_path)
                    self._send_json(200, {"ok": True, "trades": trades})
                except Exception as exc:
                    self._send_json(500, {"ok": False, "error": str(exc)})
                return
            if path == "/api/admin/accounts":
                if not self._check_admin():
                    return
                accounts = account_store.list_all()
                self._send_json(200, {"ok": True, "accounts": [a.to_dict() for a in accounts]})
                return
            if path.startswith("/api/client/"):
                token = path[len("/api/client/"):]
                self._handle_client_api(token)
                return
            if path == "/api/posts":
                posts = _get_posts(db_path)
                self._send_json(200, {"ok": True, "posts": posts})
                return
            if path == "/api/visits":
                self._send_json(200, {"ok": True, "visits": _visit_counter["count"]})
                return
            if path == "/healthz":
                self._send_json(200, {"ok": True})
                return
            if path == "/heartbeat":
                self._send_heartbeat()
                return
            if path == "/api/execution_health":
                self._send_execution_health()
                return
            self._send_json(404, {"ok": False, "error": "not found"})

        def do_POST(self):  # noqa: N802
            path = self.path.split("?", 1)[0]
            if path == "/webhook/line":
                self._handle_line_webhook()
                return
            if path == "/api/login":
                body = self._read_json()
                if body is None:
                    return
                email = (body.get("email") or "").strip().lower()
                password = (body.get("password") or "").strip()
                if not email or not password:
                    self._send_json(400, {"ok": False, "error": "Please enter your email and password."})
                    return
                acc = account_store.verify_password(email, password)
                if acc is None:
                    self._send_json(401, {"ok": False, "error": "Invalid email or password."})
                    return
                self._send_json(200, {"ok": True, "token": acc.token})
                return
            if path == "/api/fund/members":
                if not self._check_admin():
                    return
                body = self._read_json()
                if body is None:
                    return
                try:
                    m = fund_store.add_member(
                        name=body.get("name", ""),
                        email=body.get("email", ""),
                        phone=body.get("phone", ""),
                        deposit_usdt=float(body.get("deposit_usdt", 0)),
                        joined_at=body.get("joined_at", ""),
                        note=body.get("note"),
                    )
                    self._send_json(200, {"ok": True, "id": m.id})
                except Exception as exc:
                    self._send_json(500, {"ok": False, "error": str(exc)})
                return
            if path == "/api/fund/members/deactivate":
                if not self._check_admin():
                    return
                body = self._read_json()
                if body is None:
                    return
                fund_store.deactivate_member(int(body.get("id", 0)))
                self._send_json(200, {"ok": True})
                return
            if path == "/api/fund/trades":
                if not self._check_admin():
                    return
                body = self._read_json()
                if body is None:
                    return
                try:
                    t = fund_store.add_trade(
                        symbol=body.get("symbol", ""),
                        timeframe=body.get("timeframe", ""),
                        side=body.get("side", "LONG"),
                        entry=float(body.get("entry", 0)),
                        sl=float(body.get("sl", 0)),
                        opened_at=body.get("opened_at", ""),
                    )
                    self._send_json(200, {"ok": True, "id": t.id})
                except Exception as exc:
                    self._send_json(500, {"ok": False, "error": str(exc)})
                return
            if path == "/api/fund/trades/close":
                if not self._check_admin():
                    return
                body = self._read_json()
                if body is None:
                    return
                try:
                    fund_store.close_trade(
                        trade_id=int(body.get("id", 0)),
                        closed_at=body.get("closed_at", ""),
                        result=body.get("result", ""),
                        realized_rr=float(body.get("realized_rr", 0)),
                    )
                    self._send_json(200, {"ok": True})
                except Exception as exc:
                    self._send_json(500, {"ok": False, "error": str(exc)})
                return
            if path == "/api/register/step1":
                self._handle_register_step1()
                return
            if path == "/api/register/step2":
                self._handle_register_step2()
                return
            if path == "/api/posts":
                body = self._read_json()
                if body is None:
                    return
                name = (body.get("name") or "").strip()[:60]
                message = (body.get("message") or "").strip()[:2000]
                if not name or not message:
                    self._send_json(400, {"ok": False, "error": "name and message required"})
                    return
                post_id = _create_post(db_path, name, message)
                self._send_json(200, {"ok": True, "id": post_id})
                return
            if path == "/api/admin/activate":
                if not self._check_admin():
                    return
                body = self._read_json()
                if body is None:
                    return
                ok = account_store.activate(int(body.get("id", 0)))
                self._send_json(200, {"ok": ok})
                return
            if path == "/api/admin/deactivate":
                if not self._check_admin():
                    return
                body = self._read_json()
                if body is None:
                    return
                ok = account_store.deactivate(int(body.get("id", 0)))
                self._send_json(200, {"ok": ok})
                return
            if path == "/api/admin/mark_paid":
                if not self._check_admin():
                    return
                body = self._read_json()
                if body is None:
                    return
                ok = account_store.mark_paid(int(body.get("id", 0)), months=int(body.get("months", 1)))
                self._send_json(200, {"ok": ok})
                return
            if path == "/api/admin/update_api":
                if not self._check_admin():
                    return
                body = self._read_json()
                if body is None:
                    return
                api_key = (body.get("api_key") or "").strip()
                api_secret = (body.get("api_secret") or "").strip()
                if not api_key or not api_secret:
                    self._send_json(400, {"ok": False, "error": "api_key and api_secret required"})
                    return
                ok = account_store.update_api_key(int(body.get("id", 0)), api_key, api_secret)
                self._send_json(200, {"ok": ok})
                return
            self._send_json(404, {"ok": False, "error": "not found"})

        def _handle_client_api(self, token: str) -> None:
            acc = account_store.get_by_token(token)
            if acc is None:
                self._send_json(404, {"ok": False, "error": "account not found"})
                return
            if not acc.active:
                self._send_json(200, {"ok": True, "active": False, "wallet": 0, "available": 0, "upnl": 0,
                                       "positions": [], "history": [], "updated_at": ""})
                return
            try:
                from execution.binance_futures_client import BinanceFuturesClient
                client = BinanceFuturesClient(api_key=acc.api_key, api_secret=acc.api_secret)
                balance_info = client.get_account_balance()
                wallet = next((float(b["balance"]) for b in balance_info if b["asset"] == "USDT"), 0.0)
                available = next((float(b["availableBalance"]) for b in balance_info if b["asset"] == "USDT"), 0.0)
                positions_raw = client.get_position_risk()
                positions = [
                    {
                        "symbol": p["symbol"],
                        "side": "LONG" if float(p["positionAmt"]) > 0 else "SHORT",
                        "size": abs(float(p["positionAmt"])),
                        "entry_price": float(p["entryPrice"]),
                        "mark_price": float(p.get("markPrice", 0)),
                        "upnl": float(p["unRealizedProfit"]),
                        "liq_price": float(p.get("liquidationPrice", 0)),
                    }
                    for p in positions_raw if float(p.get("positionAmt", 0)) != 0
                ]
                upnl = sum(p["upnl"] for p in positions)
                history = _build_trade_history(db_path, limit=50)
                updated_at = datetime.now(UTC).strftime("%Y-%m-%d %H:%M:%S")
                self._send_json(200, {
                    "ok": True,
                    "active": True,
                    "label": acc.label,
                    "email": acc.email,
                    "wallet": round(wallet, 2),
                    "available": round(available, 2),
                    "upnl": round(upnl, 2),
                    "positions": positions,
                    "history": history,
                    "updated_at": updated_at,
                })
            except Exception as exc:
                self._send_json(500, {"ok": False, "error": str(exc)})

        def _check_admin(self) -> bool:
            ip = _get_client_ip(self)
            blocked, secs = _is_rate_limited(ip)
            if blocked:
                mins = secs // 60 + 1
                self._send_json(429, {"ok": False, "error": f"Too many failed attempts. Try again in {mins} minutes."})
                return False
            pw = self.headers.get("X-Admin-Password", "")
            if pw != _ADMIN_PASSWORD:
                _record_fail(ip)
                self._send_json(401, {"ok": False, "error": "unauthorized"})
                return False
            _record_success(ip)
            return True

        def _read_json(self) -> dict | None:
            try:
                length = int(self.headers.get("Content-Length", 0))
                body = self.rfile.read(length)
                return json.loads(body)
            except Exception:
                self._send_json(400, {"ok": False, "error": "invalid JSON"})
                return None

        def _handle_line_webhook(self) -> None:
            try:
                length = int(self.headers.get("Content-Length", 0))
                body = self.rfile.read(length)
                signature = self.headers.get("X-Line-Signature", "")
                result = handle_webhook(body, signature, account_store)
                status = result.get("status", 200)
                resp = result.get("body", "OK").encode("utf-8")
                self.send_response(status)
                self.send_header("Content-Type", "text/plain")
                self.send_header("Content-Length", str(len(resp)))
                self.end_headers()
                self.wfile.write(resp)
            except Exception as exc:
                self._send_json(500, {"ok": False, "error": str(exc)})

        def _handle_register_step1(self) -> None:
            body = self._read_json()
            if body is None:
                return
            email = (body.get("email") or "").strip().lower()
            password = (body.get("password") or "").strip()
            if not email or not password:
                self._send_json(400, {"ok": False, "error": "Please enter your email and password."})
                return
            if "@" not in email:
                self._send_json(400, {"ok": False, "error": "Invalid email address."})
                return
            if len(password) < 8:
                self._send_json(400, {"ok": False, "error": "Password must be at least 8 characters."})
                return
            if account_store.get_by_email(email):
                self._send_json(400, {"ok": False, "error": "This email is already registered."})
                return
            try:
                acc = account_store.register(email, password)
                self._send_json(200, {"ok": True, "token": acc.token})
            except Exception as exc:
                self._send_json(500, {"ok": False, "error": str(exc)})

        def _handle_register_step2(self) -> None:
            body = self._read_json()
            if body is None:
                return
            token = (body.get("token") or "").strip()
            api_key = (body.get("api_key") or "").strip()
            api_secret = (body.get("api_secret") or "").strip()
            if not token or not api_key or not api_secret:
                self._send_json(400, {"ok": False, "error": "Missing required fields."})
                return
            acc = account_store.get_by_token(token)
            if acc is None:
                self._send_json(404, {"ok": False, "error": "Account not found."})
                return
            try:
                account_store.update_api_key(acc.id, api_key, api_secret)
                try:
                    notify_new_account(acc.id, acc.email, acc.token)
                except Exception:
                    pass
                self._send_json(200, {"ok": True})
            except Exception as exc:
                self._send_json(500, {"ok": False, "error": str(exc)})

        def log_message(self, format, *args):  # noqa: A003
            return

        def _send_fund_excel(self, fstore, month: str) -> None:
            try:
                import io
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter

                data = fstore.get_settlement_data(month)
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = f"Settlement {month}"

                hdr_fill = PatternFill("solid", fgColor="1E3A5F")
                hdr_font = Font(bold=True, color="FFFFFF", size=11)
                thin = Side(style="thin", color="CCCCCC")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)

                headers = ["#", "Name", "Email", "Phone", "Deposit (USDT)", "Balance (USDT)",
                           "Trades", "Wins", "Losses", "Gross P&L", "Fee 3%", "Net P&L", "Return %"]
                ws.append(headers)
                for col, _ in enumerate(headers, 1):
                    cell = ws.cell(1, col)
                    cell.fill = hdr_fill
                    cell.font = hdr_font
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = border

                pos_font = Font(color="10B981", bold=True)
                neg_font = Font(color="EF4444", bold=True)

                for i, r in enumerate(data, 1):
                    row = [i, r["name"], r["email"], r["phone"],
                           r["initial_deposit"], r["balance_end"],
                           r["trades_count"], r["wins"], r["losses"],
                           r["gross_pnl"], r["fee"], r["net_pnl"],
                           f"{r['return_pct']}%"]
                    ws.append(row)
                    for col in range(1, len(row) + 1):
                        c = ws.cell(i + 1, col)
                        c.border = border
                        c.alignment = Alignment(horizontal="right" if col > 4 else "left")
                        if col in (10, 12) and isinstance(c.value, (int, float)):
                            c.font = pos_font if c.value >= 0 else neg_font

                # totals row
                total_row = len(data) + 2
                ws.cell(total_row, 1, "TOTAL").font = Font(bold=True)
                ws.cell(total_row, 10, sum(r["gross_pnl"] for r in data)).font = Font(bold=True)
                ws.cell(total_row, 11, sum(r["fee"] for r in data)).font = Font(bold=True)
                ws.cell(total_row, 12, sum(r["net_pnl"] for r in data)).font = Font(bold=True)

                # column widths
                widths = [4, 20, 24, 14, 14, 14, 7, 6, 7, 12, 10, 12, 10]
                for i, w in enumerate(widths, 1):
                    ws.column_dimensions[get_column_letter(i)].width = w

                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                payload = buf.read()
                fname = f"settlement_{month}.xlsx"
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="{fname}"')
                self.send_header("Content-Length", str(len(payload)))
                self.end_headers()
                self.wfile.write(payload)
            except Exception as exc:
                self._send_json(500, {"ok": False, "error": str(exc)})

        def _send_html(self, body: str) -> None:
            payload = body.encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(payload)))
            self.end_headers()
            self.wfile.write(payload)


        def _send_json(self, status: int, payload: dict) -> None:
            body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
            self.send_response(status)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Cache-Control", "no-store")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def _send_heartbeat(self) -> None:
            try:
                with open("heartbeat.txt") as f:
                    import json as _json
                    data = _json.load(f)
                self._send_json(200, {"ok": True, **data})
            except FileNotFoundError:
                self._send_json(503, {"ok": False, "error": "heartbeat.txt not found — orchestrator not running"})
            except Exception as exc:
                self._send_json(500, {"ok": False, "error": str(exc)})

        def _send_execution_health(self) -> None:
            db_path = os.getenv("WAVE_DB_PATH", "storage/wave_engine.db")
            q = ExecutionQueueStore(db_path=db_path)
            keys = [
                "execution:last_open_ok",
                "execution:last_close_ok",
                "execution:last_queue_enqueue",
                "execution:last_queue_ok",
                "execution:last_queue_error",
                "execution:circuit_opened",
                "execution:circuit_until",
                "execution:circuit_failures",
                "execution:last_de_risk_applied",
            ]
            payload = {k: read_execution_health(k, db_path=db_path) for k in keys}
            payload["queue_pending"] = q.count_pending()
            self._send_json(200, {"ok": True, "execution": payload})

    server = ThreadingHTTPServer((host, port), DashboardHandler)
    print(f"Web dashboard running at http://{host}:{port}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()
